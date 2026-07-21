from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from .decorators import role_required, student_own_data_required
from django.urls import reverse
from django.contrib import messages
from django.db.models import Q, Count
from django.utils import timezone
from django.http import JsonResponse
from datetime import date, datetime
from django.db.models import Max

try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False
from .models import (
    Student, Attendance, HostelMovement, Period, Activity, Division, Room,
    ExamType, Subject, MarkEntry, ProgressReport, AcademicYear, Enrollment,
    Grade, Section, Holiday, ExamSubjectMaxMark
)
from .forms import SectionForm, AcademicYearForm, EnquiryForm, GradeForm, DivisionForm, SubjectForm
from fees.models import FeeStructure, FeeItem


def get_holiday_dates(start_date, end_date, grade=None):
    """
    Returns a set of date objects that are holidays (Sundays, Second Saturdays,
    and manual holidays from the Holiday model) in the range [start_date, end_date].
    If grade is provided, only includes holidays that apply to that grade (or all grades).
    """
    from datetime import date, timedelta
    from students.models import Holiday
    from django.db.models import Q
    
    if not start_date or not end_date:
        return set()
        
    holidays_qs = Holiday.objects.filter(date__gte=start_date, date__lte=end_date)
    if grade:
        holidays_qs = holidays_qs.filter(Q(grades__isnull=True) | Q(grades=grade)).distinct()
        
    holiday_dates = set(holidays_qs.values_list('date', flat=True))
    
    curr = start_date
    while curr <= end_date:
        weekday = curr.weekday()
        if weekday == 6:  # Sunday
            holiday_dates.add(curr)
        elif weekday == 5:  # Saturday
            # Second Saturday check
            d = curr.day
            first_weekday = date(curr.year, curr.month, 1).weekday()
            if ((d + first_weekday) // 7 + 1) == 2:
                holiday_dates.add(curr)
        curr += timedelta(days=1)
        
    return holiday_dates


def landing_page(request):
    """Public landing page"""
    from .models import LandingPageStats
    from django.db.models import F
    
    # Increment visitor count
    stats, created = LandingPageStats.objects.get_or_create(pk=1)
    if not created:
        LandingPageStats.objects.filter(pk=1).update(visit_count=F('visit_count') + 1)
    else:
        stats.visit_count = 1
        stats.save()
        
    return render(request, 'students/landing.html')


def after_10(request):
    """Public career guide page for students after 10th grade"""
    return render(request, 'students/after_10.html')



def about_us(request):
    """Public About Us page"""
    return render(request, 'students/about_us.html')


def courses_and_fees(request):
    """Public page to view course details and link to calculators"""
    divisions = Division.objects.all().order_by('name')
    context = {
        'divisions': divisions,
        'page_title': 'Courses',
    }
    return render(request, 'students/courses_and_fees.html', context)

def course_fee_calculator(request, course_id):
    """Public page to calculate itemized fees for a specific course"""
    selected_division = get_object_or_404(Division, id=course_id)
    active_year = AcademicYear.objects.filter(is_active=True).first()
    admission_type = request.GET.get('admission_type', 'day_scholar')
    
    fee_breakdown = []
    unified_fees = []
    seen_items = set()
    
    # Get relevant grades to find general structures if needed
    relevant_grades = Grade.objects.filter(
        enrollments__division=selected_division
    ).distinct()
    
    if not relevant_grades.exists():
        relevant_grades = Grade.objects.all().order_by('order')[:3]

    # Fetch applicable fee items (Logic B Base)
    fee_items = FeeItem.objects.filter(
        Q(applicable_divisions=selected_division) | Q(applicable_divisions__isnull=True)
    ).distinct().select_related('category')

    # Fetch overriding structures (Logic A Override)
    structures = FeeStructure.objects.filter(
        academic_year=active_year,
        fee_item__in=fee_items
    ).filter(
        Q(division=selected_division) | 
        Q(division__isnull=True, grade__in=relevant_grades)
    ).select_related('fee_item')

    # Map structures by fee_item to easily override amounts
    structure_map = {}
    for fs in structures:
        # Prefer division-specific over general grade-level structure
        if fs.division == selected_division:
            structure_map[fs.fee_item.id] = fs
        elif fs.fee_item.id not in structure_map:
            structure_map[fs.fee_item.id] = fs

    for item in fee_items:
        if item.target_student_type != 'all' and item.target_student_type != admission_type:
            continue
            
        fee_key = item.name.lower().strip()
        if fee_key not in seen_items:
            seen_items.add(fee_key)
            frequency = "One Time" if item.fee_type == 'admission' else ("/ month" if item.is_monthly else "/ year")
            
            # Use structure amount if available, else default
            amount = structure_map[item.id].amount if item.id in structure_map else item.default_amount
            
            unified_fees.append({
                'name': item.name,
                'amount': amount,
                'category': item.category.name if item.category else 'General',
                'frequency': frequency
            })
        
    fee_breakdown = sorted(unified_fees, key=lambda x: x['name'])
    
    context = {
        'selected_division': selected_division,
        'fee_breakdown': fee_breakdown,
        'admission_type': admission_type,
        'page_title': f'Fee Calculator - {selected_division.name}',
        'active_year': active_year,
    }
    return render(request, 'students/course_fee_calculator.html', context)


def home(request):
    """Dashboard/Home page"""
    context = {}
    
    if request.user.is_authenticated:
        # Ensure profile exists (common for superusers created via CLI)
        if hasattr(request.user, 'profile'):
            profile = request.user.profile
        else:
            profile, created = UserProfile.objects.get_or_create(user=request.user)
            
        today = date.today()
        today_holiday = Holiday.objects.filter(date=today).first()
        context['today_holiday'] = today_holiday
        active_year = AcademicYear.objects.filter(is_active=True).first()
        context['active_year'] = active_year
        from .models import GlobalSettings
        context['suspend_student_fees'] = GlobalSettings.load().suspend_student_fees
        
        # Only calculate administrative statistics if the user is not a student
        if profile.role != 'student':
            # Statistics
            total_students = Student.objects.filter(is_active=True).count()
            hostel_students = Student.objects.filter(student_type='hostel', is_active=True).count()
            day_scholar_students = Student.objects.filter(student_type='day_scholar', is_active=True).count()
    
            # Today's attendance stats
            today_attendance = Attendance.objects.filter(date=today)
            today_present = today_attendance.filter(status='present').count()
            today_absent = today_attendance.filter(status='absent').count()
    
            # Pending hostel movements (not returned)
            pending_movements = HostelMovement.objects.filter(is_returned=False).count()
            
            # Landing page visitor count
            from .models import LandingPageStats
            stats, _ = LandingPageStats.objects.get_or_create(pk=1)
            visitor_count = stats.visit_count
    
            context.update({
                'total_students': total_students,
                'hostel_students': hostel_students,
                'day_scholar_students': day_scholar_students,
                'today_present': today_present,
                'today_absent': today_absent,
                'pending_movements': pending_movements,
                'visitor_count': visitor_count,
            })
            
        elif profile.role == 'student' and profile.student_record:
            # Student-only Portal Data
            student = profile.student_record
            st_grade = student.grade
            if st_grade:
                today_holiday = Holiday.objects.filter(date=today).filter(models.Q(grades__isnull=True) | models.Q(grades=st_grade)).first()
            else:
                today_holiday = Holiday.objects.filter(date=today).filter(grades__isnull=True).first()
            context['today_holiday'] = today_holiday
            
            # 1. Attendance Data
            student_today_att = Attendance.objects.filter(student=student, date=today).first()
            today_status = student_today_att.status if student_today_att else 'not_marked'

            # Determine effective session start date for this student's class
            enrollment = student.enrollments.filter(academic_year=context.get('active_year')).first()
            grade_obj = enrollment.grade if enrollment else st_grade
            if grade_obj and grade_obj.session_start_date:
                session_start = grade_obj.session_start_date
            elif context.get('active_year') and context['active_year'].start_date:
                session_start = context['active_year'].start_date
            else:
                session_start = None
            context['session_start'] = session_start
            
            # Monthly attendance
            import calendar
            try:
                current_month = int(request.GET.get('month', today.month))
                curr_year = int(request.GET.get('year', today.year))
                if not (1 <= current_month <= 12):
                    current_month = today.month
                if not (1900 <= curr_year <= 2100):
                    curr_year = today.year
            except (ValueError, TypeError):
                current_month = today.month
                curr_year = today.year

            monthly_att = Attendance.objects.filter(
                student=student, 
                date__year=curr_year, 
                date__month=current_month
            )
            m_start = date(curr_year, current_month, 1)
            m_end = date(curr_year, current_month, calendar.monthrange(curr_year, current_month)[1])
            m_holidays = get_holiday_dates(m_start, m_end)
            
            monthly_att_stats = monthly_att.exclude(date__in=m_holidays, status='absent')
            monthly_total = monthly_att_stats.count()
            monthly_present = monthly_att_stats.filter(status='present').count()
            monthly_attended = monthly_att_stats.filter(status__in=['present', 'late', 'excused']).count()
            attendance_monthly_percentage = round((monthly_attended / monthly_total * 100), 2) if monthly_total > 0 else 0
            current_month_name = date(curr_year, current_month, 1).strftime('%B %Y')

            # Dropdown choices
            months_choices = [(i, calendar.month_name[i]) for i in range(1, 13)]
            years_choices = list(range(today.year - 2, today.year + 2))
            
            # Yearly attendance - always from session start
            if session_start:
                yearly_att = Attendance.objects.filter(
                    student=student,
                    date__gte=session_start,
                    date__lte=active_year.end_date if active_year and active_year.end_date else today
                )
                y_end = active_year.end_date if active_year and active_year.end_date else today
                y_holidays = get_holiday_dates(session_start, y_end)
                
                yearly_att_stats = yearly_att.exclude(date__in=y_holidays, status='absent')
                yearly_total = yearly_att_stats.count()
                yearly_present = yearly_att_stats.filter(status='present').count()
                
                yearly_attended = yearly_att_stats.filter(status__in=['present', 'late', 'excused']).count()
                attendance_yearly_percentage = round((yearly_attended / yearly_total * 100), 2) if yearly_total > 0 else 0
            else:
                yearly_total = 0
                yearly_present = 0
                attendance_yearly_percentage = 0
                
            # 2. Results / Academics Data
            # All results in the active academic year
            exam_results = []
            if active_year:
                enrollment = Enrollment.objects.filter(student=student, academic_year=active_year).first()
                if enrollment and enrollment.grade:
                    # Fetch MarkEntry where exam_type's academic year is active_year
                    exam_results = MarkEntry.objects.filter(
                        student=student, 
                        enrollment__academic_year=active_year,
                        exam_type__is_published=True
                    ).select_related('exam_type', 'subject').order_by('-exam_date', 'subject__name')
            
            # Fetch ProgressReport objects for trend chart
            performance_data = []
            if active_year:
                progress_reports = ProgressReport.objects.filter(
                    student=student,
                    enrollment__academic_year=active_year,
                    exam_type__is_published=True
                ).select_related('exam_type').order_by('generated_at')
                
                for rep in progress_reports:
                    performance_data.append({
                        'exam': rep.exam_type.name,
                        'percentage': float(rep.overall_percentage)
                    })
            
            import json
            performance_data_json = json.dumps(performance_data)
            
            # 3. Hostel Movement Data
            hostel_status = None
            if student.student_type == 'hostel':
                # Check for an open 'Away' record
                open_movement = HostelMovement.objects.filter(student=student, is_returned=False).first()
                hostel_status = 'away' if open_movement else 'present'
                
            # 4. Fee Data
            student_fees = []
            fee_total_paid = 0
            fee_currently_due = 0
            fee_total_balance = 0
            try:
                from fees.models import StudentFee
                fees_list = StudentFee.objects.filter(student=student)
                for fee in fees_list:
                    student_fees.append(fee)
                    fee_total_paid += fee.amount_paid
                    if fee.balance > 0:
                        fee_total_balance += fee.balance
                        if fee.due_date and fee.due_date <= today:
                            fee_currently_due += fee.balance
            except ImportError:
                pass # If fees module is missing or disconnected
                
            context.update({
                'student_record': student,
                'today_status': today_status,
                'monthly_total': monthly_total,
                'monthly_present': monthly_present,
                'attendance_monthly_percentage': attendance_monthly_percentage,
                'current_month_name': current_month_name,
                'selected_month': current_month,
                'selected_year': curr_year,
                'months_choices': months_choices,
                'years_choices': years_choices,
                'yearly_total': yearly_total,
                'yearly_present': yearly_present,
                'attendance_yearly_percentage': attendance_yearly_percentage,
                'exam_results': exam_results,
                'performance_data_json': performance_data_json,
                'hostel_status': hostel_status,
                'fee_total_paid': fee_total_paid,
                'fee_currently_due': fee_currently_due,
                'fee_total_balance': fee_total_balance,
            })
        
    return render(request, 'students/home.html', context)


@role_required(['admin', 'teacher', 'accountant'])
def student_create(request):
    """Create a new student"""
    academic_years = AcademicYear.objects.all().order_by('-start_date')
    active_year = AcademicYear.objects.filter(is_active=True).first()
    
    if request.method == 'POST':
        student_id = request.POST.get('student_id')
        first_name = request.POST.get('first_name')
        last_name = request.POST.get('last_name')
        grade_id = request.POST.get('grade')
        division_id = request.POST.get('division') or None
        room_id = request.POST.get('room') or None
        
        section_id = request.POST.get('section') or None
        if grade_id:
            grade_obj = Grade.objects.filter(id=grade_id).first()
            if grade_obj and grade_obj.section:
                section_id = grade_obj.section.id

        student_type = request.POST.get('student_type', 'day_scholar')
        email = request.POST.get('email', '')
        phone = request.POST.get('phone', '')
        address = request.POST.get('address', '')
        bus_stop_id = request.POST.get('bus_stop') if student_type == 'day_scholar' else None
        
        form_year_id = request.POST.get('academic_year_id')
        form_year = None
        if form_year_id:
            form_year = AcademicYear.objects.filter(id=form_year_id).first()
        else:
            form_year = active_year

        # Check if student_id already exists
        if Student.objects.filter(student_id=student_id).exists():
            messages.error(request, f'Student with ID {student_id} already exists.')
        else:
            student = Student.objects.create(
                student_id=student_id,
                first_name=first_name,
                last_name=last_name,
                student_type=student_type,
                email=email,
                phone=phone,
                address=address,
                bus_stop_id=bus_stop_id,
            )
            
            siblings_ids = request.POST.getlist('siblings')
            if siblings_ids:
                student.siblings.set(siblings_ids)
            
            if form_year:
                Enrollment.objects.create(
                    student=student,
                    academic_year=form_year,
                    section_id=section_id if section_id else None,
                    grade_id=grade_id,
                    division_id=division_id if division_id else None,
                    room_id=room_id if room_id else None
                )
            
            messages.success(request, f'Student {student.full_name} created successfully for {form_year.name if form_year else "the selected year"}!')
            return redirect('students:student_list')

    divisions = Division.objects.all()
    rooms = Room.objects.all()
    sections = Section.objects.all().order_by('order', 'name')
    grades = Grade.objects.all().order_by('order', 'name')
    all_students = Student.objects.filter(is_active=True).order_by('first_name', 'last_name')
    from fees.models import BusStop
    bus_stops = BusStop.objects.all().order_by('stop_name')
    context = {
        'all_students': all_students,
        'divisions': divisions,
        'rooms': rooms,
        'sections': sections,
        'grades': grades,
        'academic_years': academic_years,
        'active_year': active_year,
        'bus_stops': bus_stops,
    }
    return render(request, 'students/student_create.html', context)


@role_required(['admin', 'accountant'])
def student_bulk_import(request):
    """Bulk import students from Excel file"""
    if not PANDAS_AVAILABLE:
        messages.error(request, 'pandas library is required for bulk import. Please install it: pip install pandas openpyxl')
        return redirect('students:student_list')

    academic_years = AcademicYear.objects.all().order_by('-start_date')
    active_year = AcademicYear.objects.filter(is_active=True).first()

    if request.method == 'POST':
        if 'excel_file' not in request.FILES:
            messages.error(request, 'Please select an Excel file.')
            return redirect('students:student_bulk_import')

        excel_file = request.FILES['excel_file']
        
        form_year_id = request.POST.get('academic_year_id')
        form_year = None
        if form_year_id:
            form_year = AcademicYear.objects.filter(id=form_year_id).first()
        else:
            form_year = active_year
            
        if not form_year:
            messages.error(request, 'No academic year selected or active.')
            return redirect('students:student_bulk_import')

        try:
            # Read Excel file
            df = pd.read_excel(excel_file)

            # Expected columns (case-insensitive matching)
            required_columns = ['student_id', 'first_name', 'last_name', 'grade']
            optional_columns = ['division', 'room', 'student_type', 'email', 'phone', 'address', 'section']

            # Normalize column names (lowercase and strip spaces)
            df.columns = df.columns.str.lower().str.strip()

            # Check required columns
            missing_columns = [col for col in required_columns if col not in df.columns]
            if missing_columns:
                messages.error(request, f'Missing required columns: {", ".join(missing_columns)}')
                return redirect('students:student_bulk_import')

            # Get divisions, rooms, and grades for lookup
            divisions = {div.name.lower(): div for div in Division.objects.all()}
            rooms = {room.room_number.lower(): room for room in Room.objects.all()}
            grades = {g.name.lower(): g for g in Grade.objects.all()}

            success_count = 0
            error_count = 0
            errors = []

            # Process each row
            for index, row in df.iterrows():
                try:
                    student_id = str(row['student_id']).strip()
                    first_name = str(row['first_name']).strip()
                    last_name = str(row['last_name']).strip()
                    grade = str(row['grade']).strip()

                    # Check if student already exists
                    if Student.objects.filter(student_id=student_id).exists():
                        errors.append(f"Row {index + 2}: Student ID {student_id} already exists")
                        error_count += 1
                        continue

                    # Validate grade is not empty
                    if not grade or grade == 'nan':
                        errors.append(f"Row {index + 2}: Grade cannot be empty")
                        error_count += 1
                        continue

                    # Get grade object
                    grade_obj = grades.get(grade.lower())
                    if not grade_obj:
                        errors.append(f"Row {index + 2}: Grade '{grade}' does not exist")
                        error_count += 1
                        continue

                    # Get division
                    division = None
                    if 'division' in df.columns and pd.notna(row.get('division')):
                        div_name = str(row['division']).strip().lower()
                        division = divisions.get(div_name)
                        if not division:
                            # Try to find by case-insensitive match
                            for div in Division.objects.all():
                                if div.name.lower() == div_name:
                                    division = div
                                    break

                    # Get room
                    room = None
                    if 'room' in df.columns and pd.notna(row.get('room')):
                        room_number = str(row['room']).strip().lower()
                        room = rooms.get(room_number)
                        if not room:
                            # Try to find by case-insensitive match
                            for r in Room.objects.all():
                                if r.room_number.lower() == room_number:
                                    room = r
                                    break

                    # Get section
                    section = None
                    if 'section' in df.columns and pd.notna(row.get('section')):
                        section_name = str(row['section']).strip().lower()
                        for s in Section.objects.all():
                            if s.name.lower() == section_name:
                                section = s
                                break

                    # Get student type
                    student_type = 'day_scholar'
                    if 'student_type' in df.columns and pd.notna(row.get('student_type')):
                        st_type = str(row['student_type']).strip().lower()
                        if st_type in ['hostel', 'day_scholar', 'day scholar']:
                            student_type = 'hostel' if st_type == 'hostel' else 'day_scholar'

                    # Get optional fields
                    email = str(row.get('email', '')).strip() if 'email' in df.columns and pd.notna(row.get('email')) else ''
                    
                    phone = str(row.get('phone', '')).strip() if 'phone' in df.columns and pd.notna(row.get('phone')) else ''
                    if phone.endswith('.0'):
                        phone = phone[:-2]
                        
                    address = str(row.get('address', '')).strip() if 'address' in df.columns and pd.notna(row.get('address')) else ''

                    # Create student
                    student = Student.objects.create(
                        student_id=student_id,
                        first_name=first_name,
                        last_name=last_name,
                        student_type=student_type,
                        email=email,
                        phone=phone,
                        address=address,
                    )
                    
                    Enrollment.objects.create(
                        student=student,
                        academic_year=form_year,
                        grade=grade_obj, # Use Grade object
                        section=section,
                        division=division,
                        room=room
                    )
                    
                    success_count += 1

                except Exception as e:
                    errors.append(f"Row {index + 2}: {str(e)}")
                    error_count += 1

            # Show results
            if success_count > 0:
                messages.success(request, f'Successfully imported {success_count} student(s) into {form_year.name}.')
            if error_count > 0:
                error_msg = f'Failed to import {error_count} student(s).'
                if errors:
                    error_msg += f' Errors: {"; ".join(errors[:5])}'  # Show first 5 errors
                    if len(errors) > 5:
                        error_msg += f' ... and {len(errors) - 5} more.'
                messages.warning(request, error_msg)

            return redirect('students:student_list')

        except Exception as e:
            messages.error(request, f'Error reading Excel file: {str(e)}')
            return redirect('students:student_bulk_import')

    # GET request - show upload form
    context = {
        'academic_years': academic_years,
        'active_year': active_year,
        'sample_columns': [
            'student_id (Required)',
            'first_name (Required)',
            'last_name (Required)',
            'grade (Required)',
            'division (Optional)',
            'room (Optional)',
            'section (Optional)',
            'student_type (Optional: day_scholar or hostel)',
            'email (Optional)',
            'phone (Optional)',
            'address (Optional)'
        ]
    }
    return render(request, 'students/student_bulk_import.html', context)


@role_required(['admin', 'accountant'])
def student_bulk_import_template(request):
    """Download an Excel template for bulk student import"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Students Import Template"

    columns = [
        ("student_id", True, "STU001", 15),
        ("first_name", True, "John", 20),
        ("last_name", True, "Doe", 20),
        ("grade", True, "10", 10),
        ("division", False, "A", 15),
        ("room", False, "101", 10),
        ("section", False, "HS", 10),
        ("student_type", False, "day_scholar", 15),
        ("email", False, "john@example.com", 25),
        ("phone", False, "1234567890", 15),
        ("address", False, "123 Main St", 30),
    ]

    # Writing headers and initial data
    for col_idx, (col_name, is_required, sample_val, width) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        # Header Style
        cell.font = Font(bold=True, color="FFFFFF")
        fill_color = "007BFF" if is_required else "6C757D" # Blue for required, grey for optional
        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        
        # Adjust width
        ws.column_dimensions[get_column_letter(col_idx)].width = width
        
        # Add sample data
        ws.cell(row=2, column=col_idx, value=sample_val)

    # Prepare HttpResponse
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="student_import_template.xlsx"'
    wb.save(response)
    
    return response

@role_required(['admin', 'accountant'])
def student_upgrade(request):
    """Upgrade students to next grade (e.g., grade 11 to 12, or 1st Year to 2nd Year)"""
    active_year = AcademicYear.objects.filter(is_active=True).first()
    if not active_year:
        messages.error(request, 'Please set an active academic year to perform upgrades.')
        return redirect('students:home')

    if request.method == 'POST':
        from_grade_name = request.POST.get('from_grade', '').strip()
        to_grade_name = request.POST.get('to_grade', '').strip()
        confirm = request.POST.get('confirm')

        if not from_grade_name or not to_grade_name:
            messages.error(request, 'Please enter both from and to grades.')
            return redirect('students:student_upgrade')

        if from_grade_name == to_grade_name:
            messages.error(request, 'From grade and to grade cannot be the same.')
            return redirect('students:student_upgrade')

        # Get Grade objects
        from_grade_obj = get_object_or_404(Grade, name=from_grade_name)
        to_grade_obj, _ = Grade.objects.get_or_create(name=to_grade_name)

        # Find latest enrollment for each active student
        enrollments_to_upgrade = []
        active_students = Student.objects.filter(is_active=True).prefetch_related('enrollments__academic_year')
        
        for student in active_students:
            # Order by academic year name descending to get latest
            enrollments = sorted(student.enrollments.all(), key=lambda x: x.academic_year.name, reverse=True)
            if enrollments:
                latest = enrollments[0]
                if latest.grade == from_grade_obj: # Compare with Grade object
                    # Also check if they already have an enrollment in the active year that is NOT from_grade
                    active_yr_enrollment = next((e for e in enrollments if e.academic_year == active_year), None)
                    if not active_yr_enrollment or active_yr_enrollment.grade == from_grade_obj: # Compare with Grade object
                        enrollments_to_upgrade.append(latest)

        count = len(enrollments_to_upgrade)

        if count == 0:
            messages.warning(request, f'No active students found in grade "{from_grade_name}" to upgrade.')
            return redirect('students:student_upgrade')

        if confirm != 'yes':
            # Show confirmation page
            context = {
                'from_grade': from_grade_name,
                'to_grade': to_grade_name,
                'student_count': count,
                'enrollments': enrollments_to_upgrade[:50],  # Show first 50 for preview
                'show_all': count > 50,
            }
            return render(request, 'students/student_upgrade_confirm.html', context)

        # Perform upgrade
        updated_count = 0
        created_count = 0
        for latest in enrollments_to_upgrade:
            # If the latest enrollment is ALREADY in the active year, we update it
            if latest.academic_year == active_year:
                latest.grade = to_grade_obj # Update with Grade object
                latest.save()
                updated_count += 1
            else:
                # Create a new enrollment for the active year
                Enrollment.objects.create(
                    student=latest.student,
                    academic_year=active_year,
                    grade=to_grade_obj, # Use Grade object
                    division=latest.division,
                    room=latest.room,
                    section=latest.section
                )
                created_count += 1

        messages.success(request, f'Successfully upgraded {updated_count + created_count} student(s) from grade "{from_grade_name}" to grade "{to_grade_name}".')
        return redirect('students:student_list')

    # GET request - show upgrade form
    # Get all existing grades from database
    existing_grades = Grade.objects.all().order_by('order', 'name') # Fetch Grade objects
    context = {
        'existing_grades': existing_grades,
    }
    return render(request, 'students/student_upgrade.html', context)


@role_required(['admin', 'accountant'])
def student_graduate(request):
    """Graduate students to Alumni (e.g., from Grade 12 or Final Year)"""
    active_year = AcademicYear.objects.filter(is_active=True).first()
    if not active_year:
        messages.error(request, 'Please set an active academic year to perform graduations.')
        return redirect('students:home')

    if request.method == 'POST':
        from_grade_name = request.POST.get('from_grade', '').strip()
        confirm = request.POST.get('confirm')

        if not from_grade_name:
            messages.error(request, 'Please select a grade to graduate.')
            return redirect('students:student_graduate')

        from_grade_obj = get_object_or_404(Grade, name=from_grade_name)

        # Find latest enrollment for each active student
        enrollments_to_graduate = []
        active_students = Student.objects.filter(is_active=True).prefetch_related('enrollments__academic_year')
        
        for student in active_students:
            # Order by academic year name descending to get latest
            enrollments = sorted(student.enrollments.all(), key=lambda x: x.academic_year.name, reverse=True)
            if enrollments:
                latest = enrollments[0]
                if latest.grade == from_grade_obj: # Compare with Grade object
                    # Only graduate if their latest enrollment is in the active year and matches from_grade
                    if latest.academic_year == active_year:
                        enrollments_to_graduate.append(latest)

        count = len(enrollments_to_graduate)

        if count == 0:
            messages.warning(request, f'No active students found in grade "{from_grade_name}" for the current academic year to graduate.')
            return redirect('students:student_graduate')

        if confirm != 'yes':
            # Show confirmation page
            context = {
                'from_grade': from_grade_name,
                'student_count': count,
                'enrollments': enrollments_to_graduate[:50],  # Show first 50 for preview
                'show_all': count > 50,
            }
            return render(request, 'students/student_graduate_confirm.html', context)

        # Perform graduation
        from alumni.models import AlumniRegistration

        graduated_count = 0
        for latest in enrollments_to_graduate:
            student = latest.student
            
            # Create Alumni record
            AlumniRegistration.objects.create(
                name=student.full_name,
                course=latest.division,
                batch=active_year.name,
                mobile_no=student.phone or ''
            )
            
            # Deactivate student
            student.is_active = False
            student.save()
            
            graduated_count += 1

        messages.success(request, f'Successfully graduated {graduated_count} student(s) from grade "{from_grade_name}" to Alumni.')
        return redirect('students:student_list')

    # GET request - show graduate form
    existing_grades = Grade.objects.all().order_by('order', 'name') # Fetch Grade objects
    context = {
        'existing_grades': existing_grades,
    }
    return render(request, 'students/student_graduate.html', context)


@role_required(['admin', 'teacher', 'accountant', 'ntstaff'])
def student_list(request):
    """List all students for the active academic year with filters"""
    if hasattr(request.user, 'profile') and request.user.profile.role == 'student':
        messages.error(request, "Access denied. Students are not allowed to view the student directory.")
        return redirect('students:home')
        
    active_year = AcademicYear.objects.filter(is_active=True).first()
    
    if not active_year:
        messages.warning(request, "Please create and set an Active Academic Year first.")
        # Fallback to just basic student list if no academic year exists
        students = Student.objects.filter(is_active=True)
    else:
        # We query Enrollments for the active year
        enrollments = Enrollment.objects.filter(academic_year=active_year, student__is_active=True).select_related('student', 'division', 'room', 'grade')

        # Filters
        grade_id = request.GET.get('grade') # Get grade ID
        division = request.GET.get('division')
        section = request.GET.get('section')
        student_type = request.GET.get('student_type')
        room = request.GET.get('room')
        search = request.GET.get('search')

        if grade_id:
            enrollments = enrollments.filter(grade_id=grade_id) # Filter by grade ID
        if division:
            enrollments = enrollments.filter(division_id=division)
        if section:
            enrollments = enrollments.filter(section_id=section)
        if student_type:
            enrollments = enrollments.filter(student__student_type=student_type)
        if room:
            enrollments = enrollments.filter(room_id=room)
        if search:
            enrollments = enrollments.filter(
                Q(student__student_id__icontains=search) |
                Q(student__first_name__icontains=search) |
                Q(student__last_name__icontains=search)
            )
        
        # Apply ordering: Grade (order then name), Division name, then Student ID
        enrollments = enrollments.order_by(
            'grade__order', 
            'grade__name', 
            'division__name', 
            'student__student_id'
        )
        
        # We will pass enrollments instead of students to the template, but we will adapt the template or rename the variable. 
        # Let's pass 'enrollments' as 'students' variable to MINIMIZE template changes initially,
        # but wait, the template expects `student.student_id`, `student.full_name`, `student.grade`. 
        # An enrollment has `enrollment.student.student_id`, `enrollment.grade`.
        # So we MUST update the template anyway to use `enrollment` objects. Let's pass 'enrollments'.
        
        students = enrollments 

    if request.GET.get('export') == 'excel':
        import openpyxl
        from django.http import HttpResponse
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="students_list.xlsx"'
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Students List"
        
        columns = ['Student ID', 'First Name', 'Last Name', 'Grade', 'Division', 'Room', 'Student Type', 'Phone', 'Email', 'Address']
        ws.append(columns)
        
        for item in students:
            if hasattr(item, 'student'): # Enrollment
                s = item.student
                grade_val = item.grade.name if item.grade else '' # Access grade name
                div_val = item.division.name if item.division else ''
                room_val = item.room.room_number if item.room else ''
            else: # Student fallback if no active year
                s = item
                grade_val = ''
                div_val = ''
                room_val = ''
                
            ws.append([
                s.student_id,
                s.first_name,
                s.last_name,
                grade_val,
                div_val,
                room_val,
                s.get_student_type_display(),
                s.phone,
                s.email,
                s.address
            ])
            
        wb.save(response)
        return response

    divisions = Division.objects.all()
    rooms = Room.objects.all()
    sections = Section.objects.all().order_by('order', 'name')
    grades = Grade.objects.all().order_by('order', 'name') # Fetch Grade objects

    context = {
        'students': students,
        'divisions': divisions,
        'rooms': rooms,
        'sections': sections,
        'grades': grades, # Pass Grade objects
        'current_filters': {
            'grade': grade_id if active_year else None, # Use grade_id
            'division': division if active_year else None,
            'section': section if active_year else None,
            'student_type': student_type if active_year else None,
            'room': room if active_year else None,
            'search': search if active_year else None,
        }
    }
    
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return render(request, 'students/partials/student_table_body.html', context)
        
    return render(request, 'students/student_list.html', context)


@role_required(['admin', 'teacher'])
def mark_attendance(request):
    """Mark attendance - Step 1: Select Date and Class"""
    attendance_type = request.GET.get('type', 'daily')
    selected_date = request.GET.get('date', date.today().isoformat())
    period_id = request.GET.get('period')
    activity_id = request.GET.get('activity')
    
    # -------------------------------------------------
    # CLASSROOM-WISE SUMMARY (GRADE + DIVISION)
    # -------------------------------------------------
    classroom_summary = OrderedDict()
    
    active_year = AcademicYear.objects.filter(is_active=True).first()
    if not active_year:
        messages.error(request, 'Please set an active academic year to mark attendance.')
        return redirect('students:home')
    
    # Get all active classes
    active_enrollments = Enrollment.objects.filter(
        academic_year=active_year, 
        student__is_active=True,
        grade__isnull=False
    ).values(
        'grade__section__id', 'grade__section__name', 
        'grade__id', 'grade__name', 
        'division__id', 'division__name'
    ).distinct()
    
    for cls in active_enrollments:
        if not cls['grade__id']: # Skip students without a grade
            continue
        g_id = cls['grade__id']
        g_name = cls['grade__name']
        d_id = cls['division__id']
        d_name = cls['division__name'] or 'No Division'
        s_id = cls['grade__section__id']
        s_name = cls['grade__section__name'] or 'No Section'
        cls_name = f"{g_name} - {d_name}"
        
        # Get count of students in this class
        student_count = Enrollment.objects.filter(academic_year=active_year, student__is_active=True, grade_id=g_id)
        if d_id:
            student_count = student_count.filter(division_id=d_id)
        else:
            student_count = student_count.filter(division__isnull=True)
            
        student_count = student_count.count()
        if student_count == 0:
            continue
            
        # Get count of recorded attendance for this class on selected date/type
        recorded_query = Attendance.objects.filter(
            date=selected_date,
            attendance_type=attendance_type,
            enrollment__grade_id=g_id,
            enrollment__academic_year=active_year
        )
        if d_id:
            recorded_query = recorded_query.filter(enrollment__division_id=d_id)
        else:
            recorded_query = recorded_query.filter(enrollment__division__isnull=True)
            
        if attendance_type == 'period' and period_id:
            recorded_query = recorded_query.filter(period_id=period_id)
        elif attendance_type == 'activity' and activity_id:
            recorded_query = recorded_query.filter(activity_id=activity_id)

        recorded_count = recorded_query.count()
        
        # Build status string (e.g., "12 / 12 Marked" or "0 / 12 Marked")
        if recorded_count == student_count:
            status = 'Complete'
            status_class = 'success'
        elif recorded_count > 0:
            status = 'Partial'
            status_class = 'warning'
        else:
            status = 'Not Marked'
            status_class = 'danger'
            
        classroom_summary[cls_name] = {
            'section_id': s_id,
            'section': s_name,
            'grade_id': g_id, # Pass grade ID
            'grade_name': g_name, # Pass grade name
            'division_id': d_id,
            'division': d_name,
            'student_count': student_count,
            'recorded_count': recorded_count,
            'status': status,
            'status_class': status_class
        }

    # Sorting
    division_summary = {k: v for k, v in classroom_summary.items() if v['division'] != 'No Division'}

    periods = Period.objects.all()
    activities = Activity.objects.filter(date=selected_date) if selected_date else Activity.objects.none()

    context = {
        'attendance_type': attendance_type,
        'selected_date': selected_date,
        'periods': periods,
        'activities': activities,
        'selected_period': period_id,
        'selected_activity': activity_id,
        'classroom_summary': classroom_summary,
        'division_summary': division_summary,
    }
    return render(request, 'students/mark_attendance.html', context)

@role_required(['admin', 'teacher'])
def mark_attendance_class(request, grade_id, division_id): # Changed grade to grade_id
    """Mark attendance - Step 2: Enter attendance for specific class"""
    attendance_type = request.GET.get('type', 'daily')
    selected_date = request.GET.get('date', date.today().isoformat())
    period_id = request.GET.get('period')
    activity_id = request.GET.get('activity')
    section_id = request.GET.get('section')
    
    # Handle division matching 'None' (from URL)
    actual_division_id = None if division_id == 0 else division_id
    
    active_year = AcademicYear.objects.filter(is_active=True).first()
    if not active_year:
        messages.error(request, 'Please set an active academic year to mark attendance.')
        return redirect('students:home')

    grade_obj = get_object_or_404(Grade, id=grade_id) # Get Grade object

    # Get enrollments based on class
    enrollments = Enrollment.objects.filter(academic_year=active_year, student__is_active=True, grade=grade_obj).select_related('student', 'division', 'room', 'section')
    
    if grade_obj.section:
        section_name = grade_obj.section.name
    else:
        section_name = "No Section"
        
    if actual_division_id:
        enrollments = enrollments.filter(division_id=actual_division_id)
        division = get_object_or_404(Division, id=actual_division_id)
        class_name = f"{section_name} - {grade_obj.name} - {division.name}" 
    else:
        enrollments = enrollments.filter(division__isnull=True)
        class_name = f"{section_name} - {grade_obj.name} - No Division"

    # Get existing attendance for the date
    existing_attendance = {}
    if attendance_type == 'period' and period_id:
        existing_attendance = {
            att.enrollment_id: att.status
            for att in Attendance.objects.filter(
                date=selected_date,
                attendance_type='period',
                period_id=period_id,
                enrollment__in=enrollments
            )
        }
    elif attendance_type == 'activity' and activity_id:
        existing_attendance = {
            att.enrollment_id: att.status
            for att in Attendance.objects.filter(
                date=selected_date,
                attendance_type='activity',
                activity_id=activity_id,
                enrollment__in=enrollments
            )
        }
    elif attendance_type == 'daily':
        existing_attendance = {
            att.enrollment_id: att.status
            for att in Attendance.objects.filter(
                date=selected_date,
                attendance_type='daily',
                enrollment__in=enrollments
            )
        }

    if request.method == 'POST':
        # Automatically determine who is marking the attendance from the logged-in user
        marked_by = request.user.get_full_name() or request.user.username
        if not marked_by:
            marked_by = "Admin" # Fallback if user somehow has no name/username
            
        attendance_data = request.POST.getlist('attendance')
        success_count = 0

        for data in attendance_data:
            enrollment_id, status = data.split('|')
            enrollment = get_object_or_404(Enrollment, id=enrollment_id)
            student = enrollment.student

            # Build filter for finding existing attendance
            filter_kwargs = {
                'student': student,
                'enrollment': enrollment,
                'date': selected_date,
                'attendance_type': attendance_type,
            }

            if attendance_type == 'period':
                if period_id:
                    filter_kwargs['period_id'] = period_id
                else:
                    continue  # Skip if period is required but not provided
            elif attendance_type == 'activity':
                if activity_id:
                    filter_kwargs['activity_id'] = activity_id
                else:
                    continue  # Skip if activity is required but not provided
            else:  # daily
                filter_kwargs['period__isnull'] = True
                filter_kwargs['activity__isnull'] = True

            attendance, created = Attendance.objects.update_or_create(
                defaults={
                    'status': status,
                    'marked_by': marked_by,
                },
                **filter_kwargs
            )
            success_count += 1

        params = [f"type={attendance_type}", f"date={selected_date}"]
        if period_id: params.append(f"period={period_id}")
        if activity_id: params.append(f"activity={activity_id}")
        
        return redirect(f"{reverse('students:mark_attendance')}?{'&'.join(params)}")

    context = {
        'class_name': class_name,
        'grade_id': grade_id, # Pass grade ID
        'grade_name': grade_obj.name, # Pass grade name
        'division_id': division_id,
        'section_id': section_id,
        'enrollments': enrollments,
        'attendance_type': attendance_type,
        'selected_date': selected_date,
        'existing_attendance': existing_attendance,
        'selected_period': period_id,
        'selected_activity': activity_id,
    }
    return render(request, 'students/mark_attendance_class.html', context)


from collections import OrderedDict
from django.db.models import Count, Q
from django.shortcuts import render
from django.utils.timezone import localtime
import calendar
from datetime import date
from django.db.models import Exists, OuterRef


from collections import OrderedDict
from datetime import date
import calendar

from django.shortcuts import render
from django.db.models import Count, Q
from django.utils.timezone import localtime

from .models import Attendance, Grade
from students.models import Student, Division



@role_required(['admin', 'teacher'])
def attendance_list(request):
    """
    Attendance List with:
    - Advanced filters
    - Division-wise summary
    - Monthly Class + Division audit table
    - Sunday & Second Saturday holiday marking
    """

    # -------------------------------------------------
    # GET FILTERS
    # -------------------------------------------------
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    attendance_type = request.GET.get('attendance_type')
    status = request.GET.get('status')
    section = request.GET.get('section')
    grade_id = request.GET.get('grade') # Get grade ID
    division = request.GET.get('division')
    student_id = request.GET.get('student_id')

    # Month / Year for audit
    today = date.today()
    month = int(request.GET.get('month', today.month))
    year = int(request.GET.get('year', today.year))

    active_year = AcademicYear.objects.filter(is_active=True).first()

    # -------------------------------------------------
    # BASE QUERYSET
    # -------------------------------------------------
    attendances = (
        Attendance.objects
        .filter(enrollment__academic_year=active_year)
        .select_related('student', 'enrollment', 'enrollment__division', 'enrollment__grade', 'period', 'activity')
        .order_by('-date', 'enrollment__grade__order', 'enrollment__grade__name', 'student__last_name') # Order by grade object
    )

    # -------------------------------------------------
    # APPLY FILTERS
    # -------------------------------------------------
    if date_from:
        attendances = attendances.filter(date__gte=date_from)

    if date_to:
        attendances = attendances.filter(date__lte=date_to)

    if attendance_type:
        attendances = attendances.filter(attendance_type=attendance_type)

    if status:
        attendances = attendances.filter(status=status)

    if grade_id:
        attendances = attendances.filter(enrollment__grade_id=grade_id) # Filter by grade ID

    if division:
        attendances = attendances.filter(enrollment__division_id=division)
        
    if section:
        attendances = attendances.filter(enrollment__section_id=section)

    if student_id:
        attendances = attendances.filter(
            student__student_id__icontains=student_id
        )

    # -------------------------------------------------
    # CLASSROOM-WISE SUMMARY (GRADE + DIVISION)
    # -------------------------------------------------
    classroom_summary = OrderedDict()
    
    # Pre-populate with all expected classes so they show even if no attendance
    active_enrollments = Enrollment.objects.filter(academic_year=active_year, student__is_active=True).values('section__id', 'section__name', 'grade__id', 'grade__name', 'division__id', 'division__name').distinct()
    for student_cls in active_enrollments:
        if not student_cls['grade__id']: # Skip students without a grade
            continue
        g_id = student_cls['grade__id']
        g_name = student_cls['grade__name']
        d_id = student_cls['division__id']
        d_name = student_cls['division__name'] or 'No Division'
        s_id = student_cls['section__id']
        s_name = student_cls['section__name'] or 'No Section'
        cls_name = f"{s_name} - {g_name} - {d_name}"
        
        classroom_summary[cls_name] = {
            'section_id': s_id,
            'section': s_name,
            'grade_id': g_id,
            'grade_name': g_name,
            'division_id': d_id,
            'division': d_name,
            'present': 0,
            'absent': 0,
            'late': 0,
            'excused': 0,
            'total': 0,
            'last_marker': None,
            'last_updated': None,
            'today_stats': None,
        }

    stats_attendances = attendances
    if attendances.exists():
        dates = list(attendances.values_list('date', flat=True))
        if dates:
            h_dates = get_holiday_dates(min(dates), max(dates))
            stats_attendances = attendances.exclude(date__in=h_dates, status='absent')
            
    if stats_attendances.exists():
        classroom_stats = (
            stats_attendances
            .values(
                'enrollment__section__id',
                'enrollment__section__name',
                'enrollment__grade__id',
                'enrollment__grade__name',
                'enrollment__division__id',
                'enrollment__division__name'
            )
            .annotate(
                total=Count('id'),
                present=Count('id', filter=Q(status='present')),
                absent=Count('id', filter=Q(status='absent')),
                late=Count('id', filter=Q(status='late')),
                excused=Count('id', filter=Q(status='excused')),
            )
            .order_by('enrollment__grade__order', 'enrollment__grade__name', 'enrollment__division__name')
        )

        for stat in classroom_stats:
            grade_id_val = stat['enrollment__grade__id']
            grade_name_val = stat['enrollment__grade__name']
            div_id = stat['enrollment__division__id']
            div_name = stat['enrollment__division__name'] or 'No Division'
            sec_id = stat['enrollment__section__id']
            sec_name = stat['enrollment__section__name'] or 'No Section'

            classroom_name = f"{sec_name} - {grade_name_val} - {div_name}"

            latest_entry = (
                attendances
                .filter(
                    enrollment__grade_id=grade_id_val,
                    enrollment__division_id=div_id
                )
                .order_by('-updated_at', '-created_at', '-id')
                .first()
            )

            last_marker = getattr(latest_entry, 'marked_by', None)
            last_updated = None

            if latest_entry:
                last_dt = (
                    getattr(latest_entry, 'updated_at', None)
                    or getattr(latest_entry, 'created_at', None)
                )
                if last_dt:
                    last_updated = localtime(last_dt)

            if classroom_name in classroom_summary:
                classroom_summary[classroom_name].update({
                    'present': stat['present'],
                    'absent': stat['absent'],
                    'late': stat['late'],
                    'excused': stat['excused'],
                    'total': stat['total'],
                    'last_marker': last_marker,
                    'last_updated': last_updated,
                })
            else:
                classroom_summary[classroom_name] = {
                    'section_id': sec_id,
                    'section': sec_name,
                    'grade_id': grade_id_val,
                    'grade_name': grade_name_val,
                    'division_id': div_id,
                    'division': div_name,
                    'present': stat['present'],
                    'absent': stat['absent'],
                    'late': stat['late'],
                    'excused': stat['excused'],
                    'total': stat['total'],
                    'last_marker': last_marker,
                    'last_updated': last_updated,
                    'today_stats': None,
                }
                
    # -------------------------------------------------
    # TODAY'S CLASSROOM-WISE SUMMARY (GRADE + DIVISION)
    # -------------------------------------------------
    
    updated_classrooms = []
    not_updated_classrooms = []
    
    
    today_attendances = Attendance.objects.filter(date=today, enrollment__academic_year=active_year).select_related('student', 'enrollment', 'enrollment__division', 'enrollment__grade')
    
    if today_attendances.exists():
        today_stats = (
            today_attendances
            .values(
                'enrollment__section__name',
                'enrollment__grade__name',
                'enrollment__division__id',
                'enrollment__division__name'
            )
            .annotate(
                total=Count('id'),
                present=Count('id', filter=Q(status='present')),
                absent=Count('id', filter=Q(status='absent')),
                late=Count('id', filter=Q(status='late')),
                excused=Count('id', filter=Q(status='excused')),
            )
        )
        for stat in today_stats:
            grade_name_val = stat['enrollment__grade__name']
            div_name_val = stat['enrollment__division__name'] or 'No Division'
            sec_name_val = stat['enrollment__section__name'] or 'No Section'
            cls_name = f"{sec_name_val} - {grade_name_val} - {div_name_val}"
            
            if cls_name in classroom_summary:
                classroom_summary[cls_name]['today_stats'] = {
                    'total': stat['total'],
                    'present': stat['present'],
                    'absent': stat['absent'],
                    'late': stat['late'],
                    'excused': stat['excused'],
                }

    # Now populate updated/not_updated lists based on today_stats
    for cls_name, info in classroom_summary.items():
        if info['today_stats']:
            updated_classrooms.append(cls_name)
        else:
            not_updated_classrooms.append(cls_name)

    # -------------------------------------------------
    # RENDER CONTEXT
    # -------------------------------------------------
    
    # Exclude students with division 'None' from division_wise summary
    division_summary = {k: v for k, v in classroom_summary.items() if v['division'] != 'No Division'}

    context = {
        'attendances': attendances,
        'classroom_summary': classroom_summary,
        'division_summary': division_summary,
        'updated_classrooms': updated_classrooms,
        'not_updated_classrooms': not_updated_classrooms,

        # For the drop-downs
        'grades': Grade.objects.all().order_by('order', 'name'), # Fetch Grade objects
        'divisions': Division.objects.all(),
        'sections': Section.objects.all().order_by('order', 'name'),
        
        # State
        'current_filters': {
            'date_from': date_from,
            'date_to': date_to,
            'attendance_type': attendance_type,
            'status': status,
            'student_id': student_id,
            'section': section,
            'grade': grade_id, # Use grade_id
            'division': division
        },
    }

    return render(request, 'students/attendance_list.html', context)

@role_required(['admin', 'teacher'])
def today_attendance_view(request):
    active_year = AcademicYear.objects.filter(is_active=True).first()
    if not active_year:
        messages.error(request, 'Please set an active academic year to view attendance.')
        return redirect('students:home')

    today = date.today()
    today_holiday = Holiday.objects.filter(date=today).first()
    enrollments = Enrollment.objects.filter(
        academic_year=active_year, 
        student__is_active=True
    ).select_related('student', 'grade', 'division', 'section')

    attendances = Attendance.objects.filter(date=today, attendance_type='daily')
    att_map = {att.student_id: att.status for att in attendances}

    status_display_map = {
        'present': 'Present',
        'absent': 'Absent',
        'late': 'Late',
        'excused': 'Excused',
        'not_marked': 'Not Marked',
        'not marked': 'Not Marked',
    }
    student_stats = []
    for env in enrollments:
        raw_status = att_map.get(env.student_id, 'not_marked')
        status = raw_status.replace(' ', '_')
        env_grade = env.grade
        student_holiday = Holiday.objects.filter(date=today).filter(models.Q(grades__isnull=True) | models.Q(grades=env_grade)).first() if env_grade else Holiday.objects.filter(date=today, grades__isnull=True).first()

        status_disp = status_display_map.get(status, status_display_map.get(raw_status, 'Not Marked'))
        if student_holiday and raw_status == 'not_marked':
            status = 'holiday'
            status_disp = f"Holiday ({student_holiday.title})"

        student_stats.append({
            'enrollment': env,
            'student': env.student,
            'status': status,
            'status_display': status_disp,
        })

    # Custom Sorting: First 11, then 12, then Degree, then others
    def get_today_attendance_sort_key(stat):
        env = stat['enrollment']
        g_name = (env.grade.name.upper() if env.grade else '').strip()
        sec_name = (env.section.name.upper() if env.section else (env.grade.section.name.upper() if env.grade and env.grade.section else '')).strip()

        if '11' in g_name:
            cat = 1
        elif '12' in g_name:
            cat = 2
        elif 'DEGREE' in sec_name or 'YEAR' in g_name or 'DEGREE' in g_name:
            cat = 3
        else:
            cat = 4

        g_order = getattr(env.grade, 'order', 999) if env.grade else 999
        d_name = env.division.name if env.division else 'ZZZ'
        s_first = env.student.first_name if env.student else ''
        s_last = env.student.last_name if env.student else ''

        return (cat, g_order, g_name, d_name, s_first, s_last)

    student_stats.sort(key=get_today_attendance_sort_key)

    present_count = len([s for s in student_stats if s['status'] == 'present'])
    absent_count = len([s for s in student_stats if s['status'] == 'absent'])
    late_count = len([s for s in student_stats if s['status'] == 'late'])
    excused_count = len([s for s in student_stats if s['status'] == 'excused'])
    not_marked_count = len([s for s in student_stats if s['status'] == 'not_marked'])
    total_present_all = present_count + late_count + excused_count

    context = {
        'today': today,
        'student_stats': student_stats,
        'present_count': present_count,
        'absent_count': absent_count,
        'late_count': late_count,
        'excused_count': excused_count,
        'not_marked_count': not_marked_count,
        'total_present_all': total_present_all,
        'total_count': len(student_stats),
        'today_holiday': today_holiday,
    }
    return render(request, 'students/today_attendance.html', context)

@role_required(['admin', 'ntstaff'])
def hostel_movement_list(request):
    """List hostel movements with stats and enhanced search"""
    from django.db.models import Q
    movements = HostelMovement.objects.select_related('student')

    # Calculate Stats
    total_hostel_students = Student.objects.filter(student_type='hostel', is_active=True).count()
    # Away students are those with a movement log where is_returned is False
    away_students = HostelMovement.objects.filter(is_returned=False, student__is_active=True).values('student').distinct().count()
    present_students = total_hostel_students - away_students

    # Filters
    student_id = request.GET.get('student_id')
    is_returned = request.GET.get('is_returned')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')

    if student_id:
        movements = movements.filter(
            Q(student__student_id__icontains=student_id) | 
            Q(student__first_name__icontains=student_id) |
            Q(student__last_name__icontains=student_id)
        )
    if is_returned is not None:
        movements = movements.filter(is_returned=is_returned == '1')
    if date_from:
        movements = movements.filter(departure_date__gte=date_from)
    if date_to:
        movements = movements.filter(departure_date__lte=date_to)

    movements = movements.order_by('-departure_date', '-departure_time')

    context = {
        'movements': movements,
        'stats': {
            'total': total_hostel_students,
            'present': present_students,
            'away': away_students
        },
        'current_filters': {
            'student_id': student_id,
            'is_returned': is_returned,
            'date_from': date_from,
            'date_to': date_to,
        }
    }
    return render(request, 'students/hostel_movement_list.html', context)


@role_required(['admin', 'ntstaff'])
def hostel_movement_create(request):
    """Create hostel movement record"""
    if request.method == 'POST':
        student_id = request.POST.get('student')
        departure_date = request.POST.get('departure_date')
        departure_time = request.POST.get('departure_time')
        escorting_person = request.POST.get('escorting_person')
        reason = request.POST.get('reason')
        expected_return_date = request.POST.get('expected_return_date') or None

        student = get_object_or_404(Student, id=student_id, student_type='hostel')

        movement = HostelMovement.objects.create(
            student=student,
            departure_date=departure_date,
            departure_time=departure_time,
            escorting_person=escorting_person,
            reason=reason,
            expected_return_date=expected_return_date,
        )

        messages.success(request, f'Movement record created for {student.full_name}')
        return redirect('students:hostel_movement_list')

    active_year = AcademicYear.objects.filter(is_active=True).first()
    enrollments = Enrollment.objects.filter(
        academic_year=active_year,
        student__student_type='hostel', 
        student__is_active=True
    ).select_related('student')
    context = {'enrollments': enrollments}
    return render(request, 'students/hostel_movement_create.html', context)


@role_required(['admin', 'ntstaff'])
def hostel_movement_update(request, pk):
    """Update hostel movement (mark as returned)"""
    movement = get_object_or_404(HostelMovement, pk=pk)

    if request.method == 'POST':
        arrival_date = request.POST.get('arrival_date')
        arrival_time = request.POST.get('arrival_time')
        sign = request.POST.get('sign')
        remarks = request.POST.get('remarks')
        is_returned = request.POST.get('is_returned') == 'on'

        movement.arrival_date = arrival_date if arrival_date else None
        movement.arrival_time = arrival_time if arrival_time else None
        movement.sign = sign
        movement.remarks = remarks
        movement.is_returned = is_returned
        movement.save()

        messages.success(request, f'Movement record updated for {movement.student.full_name}')
        return redirect('students:hostel_movement_list')

    context = {'movement': movement}
    return render(request, 'students/hostel_movement_update.html', context)


@role_required(['admin', 'teacher'])
def exam_type_list(request):
    """List all exam types"""
    exam_types = ExamType.objects.all()
    context = {
        'exam_types': exam_types,
    }
    return render(request, 'students/exam_type_list.html', context)


@role_required(['admin', 'teacher'])
def exam_type_create(request):
    """Create a new exam type"""
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        description = request.POST.get('description', '').strip()
        subject_type = request.POST.get('subject_type', 'all')
        section_id = request.POST.get('section')
        order = request.POST.get('order', 0)
        is_published = request.POST.get('is_published') == 'on'
        subject_ids = request.POST.getlist('subjects')

        if not name:
            messages.error(request, 'Exam type name is required.')
            return redirect('students:exam_type_create')

        # Check if exam type already exists
        if ExamType.objects.filter(name__iexact=name).exists():
            messages.error(request, f'Exam type "{name}" already exists.')
            return redirect('students:exam_type_create')

        try:
            order = int(order) if order else 0
        except:
            order = 0

        section = None
        if section_id:
            section = Section.objects.filter(id=section_id).first()

        exam_type = ExamType.objects.create(
            name=name,
            description=description,
            subject_type=subject_type,
            section=section,
            order=order,
            is_published=is_published,
        )
        exam_type.subjects.set(subject_ids)
        
        messages.success(request, f'Exam type "{exam_type.name}" created successfully!')
        return redirect('students:exam_type_list')

    # Pass the choices to the template
    sections = Section.objects.all().order_by('order', 'name')
    subjects = Subject.objects.filter(is_active=True).select_related('grade', 'division', 'section').order_by('grade__order', 'grade__name', 'subject_type', 'name')
    context = {
        'subject_type_choices': ExamType.SUBJECT_TYPE_CHOICES,
        'sections': sections,
        'subjects': subjects,
    }
    return render(request, 'students/exam_type_create.html', context)


@role_required(['admin', 'teacher'])
def exam_type_update(request, pk):
    """Update an exam type"""
    exam_type = get_object_or_404(ExamType, pk=pk)

    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        description = request.POST.get('description', '').strip()
        subject_type = request.POST.get('subject_type', 'all')
        section_id = request.POST.get('section')
        order = request.POST.get('order', 0)
        is_published = request.POST.get('is_published') == 'on'
        subject_ids = request.POST.getlist('subjects')

        if not name:
            messages.error(request, 'Exam type name is required.')
            return redirect('students:exam_type_update', pk=pk)

        # Check if name already exists (excluding current)
        if ExamType.objects.filter(name__iexact=name).exclude(pk=pk).exists():
            messages.error(request, f'Exam type "{name}" already exists.')
            return redirect('students:exam_type_update', pk=pk)

        try:
            order = int(order) if order else 0
        except:
            order = exam_type.order

        section = None
        if section_id:
            section = Section.objects.filter(id=section_id).first()

        exam_type.name = name
        exam_type.description = description
        exam_type.subject_type = subject_type
        exam_type.section = section
        exam_type.order = order
        exam_type.is_published = is_published
        exam_type.save()
        exam_type.subjects.set(subject_ids)

        messages.success(request, f'Exam type "{exam_type.name}" updated successfully!')
        return redirect('students:exam_type_list')

    sections = Section.objects.all().order_by('order', 'name')
    subjects = Subject.objects.filter(is_active=True).select_related('grade', 'division', 'section').order_by('grade__order', 'grade__name', 'subject_type', 'name')
    selected_subject_ids = set(exam_type.subjects.values_list('id', flat=True))
    context = {
        'exam_type': exam_type,
        'subject_type_choices': ExamType.SUBJECT_TYPE_CHOICES,
        'sections': sections,
        'subjects': subjects,
        'selected_subject_ids': selected_subject_ids,
    }
    return render(request, 'students/exam_type_update.html', context)


@role_required(['admin', 'teacher'])
def exam_type_delete(request, pk):
    """Delete an exam type"""
    exam_type = get_object_or_404(ExamType, pk=pk)

    if request.method == 'POST':
        name = exam_type.name
        exam_type.delete()
        messages.success(request, f'Exam type "{name}" deleted successfully!')
        return redirect('students:exam_type_list')

    context = {
        'exam_type': exam_type,
    }
    return render(request, 'students/exam_type_delete.html', context)


@role_required(['admin', 'teacher'])
def exam_type_toggle_publish(request, pk):
    """Toggle publication status of an exam"""
    exam_type = get_object_or_404(ExamType, pk=pk)
    exam_type.is_published = not exam_type.is_published
    exam_type.save()
    status = "published" if exam_type.is_published else "unpublished"
    messages.success(request, f'Exam "{exam_type.name}" is now {status}.')
    return redirect('students:exam_type_list')


@role_required(['admin', 'teacher'])
def exam_subject_maxmarks(request, exam_type_id):
    """Manage per-exam maximum marks for each subject.
    Displays all subjects grouped by grade; allows staff to set exam-specific max marks.
    Falls back to Subject.max_marks if no override is saved."""
    exam_type = get_object_or_404(ExamType, pk=exam_type_id)

    # Optional grade filter
    grade_id = request.GET.get('grade_id')
    grade_obj = None
    if grade_id:
        grade_obj = Grade.objects.filter(pk=grade_id).first()

    # Build subject queryset
    subjects_qs = Subject.objects.filter(is_active=True).select_related('grade', 'division', 'section')
    if exam_type.section:
        subjects_qs = subjects_qs.filter(Q(section=exam_type.section) | Q(grade__section=exam_type.section))
    if exam_type.subjects.exists():
        subjects_qs = subjects_qs.filter(id__in=exam_type.subjects.all())
    if grade_obj:
        subjects_qs = subjects_qs.filter(grade=grade_obj)
    subjects = list(subjects_qs.order_by('grade__order', 'grade__name', 'subject_type', 'name'))

    if request.method == 'POST':
        saved = 0
        for subject in Subject.objects.filter(is_active=True):
            field_name = f'max_marks_{subject.id}'
            val_str = request.POST.get(field_name, '').strip()
            if val_str == '':
                # No value submitted — remove any existing override (revert to subject default)
                ExamSubjectMaxMark.objects.filter(exam_type=exam_type, subject=subject).delete()
            else:
                try:
                    val = int(float(val_str))
                    if val > 0:
                        ExamSubjectMaxMark.objects.update_or_create(
                            exam_type=exam_type,
                            subject=subject,
                            defaults={'max_marks': val}
                        )
                        saved += 1
                    else:
                        ExamSubjectMaxMark.objects.filter(exam_type=exam_type, subject=subject).delete()
                except (ValueError, TypeError):
                    pass
        messages.success(request, f'Saved max marks for {saved} subject(s) for "{exam_type.name}".')
        redirect_url = reverse('students:exam_subject_maxmarks', args=[exam_type_id])
        if grade_id:
            redirect_url += f'?grade_id={grade_id}'
        return redirect(redirect_url)

    # Build existing overrides lookup: {subject_id: max_marks}
    overrides = {
        esm.subject_id: esm.max_marks
        for esm in ExamSubjectMaxMark.objects.filter(exam_type=exam_type)
    }

    # Annotate subjects with their effective max marks
    for s in subjects:
        s.effective_max = overrides.get(s.id, s.max_marks)
        s.has_override  = s.id in overrides

    # All available grades (for the filter dropdown)
    grades_qs = Grade.objects.filter(subjects__is_active=True).distinct().order_by('order', 'name')
    if exam_type.section:
        grades_qs = grades_qs.filter(section=exam_type.section)
    grades = list(grades_qs)

    context = {
        'exam_type': exam_type,
        'subjects': subjects,
        'grades': grades,
        'selected_grade': grade_obj,
    }
    return render(request, 'students/exam_subject_maxmarks.html', context)


@role_required(['admin', 'teacher'])
def mark_entry_step1(request):
    """Step 1: Select Exam Type"""
    exam_types = ExamType.objects.all().order_by('order', 'name')
    return render(request, 'students/mark_entry_step1.html', {'exam_types': exam_types})

@role_required(['admin', 'teacher'])
def mark_entry_step2(request, exam_type_id):
    """Step 2: Select Class (Combination of Grade and Division)"""
    exam_type = get_object_or_404(ExamType, id=exam_type_id)
    
    # Get all active students to find unique Grade + Division combinations through Enrollment
    active_year = AcademicYear.objects.filter(is_active=True).first()
    enrollments_query = Enrollment.objects.filter(academic_year=active_year, student__is_active=True).select_related('division', 'grade')
    if exam_type.section:
        enrollments_query = enrollments_query.filter(section=exam_type.section)
    enrollments = enrollments_query
    
    # Extract unique classes
    classes_dict = {}
    for enrollment in enrollments:
        grade_obj = enrollment.grade
        div_id = enrollment.division_id
        div_name = enrollment.division.name if enrollment.division else None
        
        if not grade_obj:
            continue
        
        # Create a unique key for the combination
        key = f"{grade_obj.id}_{div_id or 'none'}"
        
        if key not in classes_dict:
            classes_dict[key] = {
                'grade_id': grade_obj.id,
                'grade_name': grade_obj.name,
                'division_id': div_id,
                'division_name': div_name,
                'student_count': 0
            }
        
        classes_dict[key]['student_count'] += 1
        
    # Sort classes logically (numeric grades first, then by division)
    def class_sort_key(c):
        grade_name = c['grade_name']
        grade_val = (0, int(grade_name)) if grade_name.isdigit() else (1, grade_name.lower())
        div_val = c['division_name'].lower() if c['division_name'] else ''
        return (grade_val, div_val)
        
    classes = sorted(classes_dict.values(), key=class_sort_key)
    
    return render(request, 'students/mark_entry_step2.html', {
        'exam_type': exam_type,
        'classes': classes
    })

@role_required(['admin', 'teacher'])
def mark_entry_step3(request, exam_type_id):
    """Step 3: Enter marks for all subjects for the selected class"""
    exam_type = get_object_or_404(ExamType, id=exam_type_id)
    grade_id = request.GET.get('grade_id') # Get grade ID
    division_id = request.GET.get('division_id')
    
    if not grade_id:
        messages.error(request, 'Grade is required.')
        return redirect('students:mark_entry_step2', exam_type_id=exam_type.id)
        
    grade_obj = get_object_or_404(Grade, id=grade_id) # Get Grade object
    
    # Get students for this class via Enrollment
    active_year = AcademicYear.objects.filter(is_active=True).first()
    enrollments_query = Enrollment.objects.filter(academic_year=active_year, student__is_active=True, grade=grade_obj).select_related('student')
    if exam_type.section:
        enrollments_query = enrollments_query.filter(section=exam_type.section)
        
    if division_id and division_id != 'None':
        enrollments_query = enrollments_query.filter(division_id=division_id)
        division_name = Division.objects.get(id=division_id).name
    else:
        enrollments_query = enrollments_query.filter(division__isnull=True)
        division_id = None
        division_name = None
        
    enrollments = list(enrollments_query.order_by('student__last_name', 'student__first_name'))
    students = [e.student for e in enrollments]
    
    if not students:
        messages.error(request, 'No active students found for this class.')
        return redirect('students:mark_entry_step2', exam_type_id=exam_type.id)
        
    # Get subjects for this class
    subjects_query = Subject.objects.filter(is_active=True, grade=grade_obj) # Filter by Grade object
    if exam_type.subjects.exists():
        subjects_query = subjects_query.filter(id__in=exam_type.subjects.all())
    
    # First, filter by the exam's subject_type preference
    if exam_type.subject_type != 'all':
        subjects_query = subjects_query.filter(subject_type=exam_type.subject_type)
        
    # Then filter by division
    if division_id:
        if exam_type.subject_type == 'hadiya':
            # If specifically asking for hadiya, just ensure no division constraint blocks it
            subjects_query = subjects_query.filter(division__isnull=True)
        elif exam_type.subject_type == 'division':
            # If specifically asking for division subjects, must match the division
            subjects_query = subjects_query.filter(division_id=division_id)
        else:
            # 'all' - Show both division subjects and common (hadiya) subjects
            subjects_query = subjects_query.filter(
                Q(division_id=division_id) | Q(subject_type='hadiya', division__isnull=True)
            )
    else:
        # If no division, only show common/Hadiya subjects or subjects with no division
        subjects_query = subjects_query.filter(division__isnull=True)
        
    all_class_subjects = list(subjects_query.order_by('subject_type', 'name'))
    subjects = all_class_subjects

    # ── Annotate subjects with exam-specific max marks ──────────────────────────
    # Load ExamSubjectMaxMark overrides for this exam, for all subjects in this class
    exam_overrides = {
        esm.subject_id: esm.max_marks
        for esm in ExamSubjectMaxMark.objects.filter(
            exam_type=exam_type,
            subject__in=all_class_subjects
        )
    }
    for s in all_class_subjects:
        s.effective_max_marks = exam_overrides.get(s.id, s.max_marks)
    # ────────────────────────────────────────────────────────────────────────────

    selected_subject_id = request.GET.get('subject_id')
    selected_subject = None
    if selected_subject_id and selected_subject_id != 'all':
        try:
            selected_subject_id_int = int(selected_subject_id)
            selected_subject = next((s for s in all_class_subjects if s.id == selected_subject_id_int), None)
            if selected_subject:
                subjects = [selected_subject]
        except (ValueError, TypeError):
            pass
    
    if request.method == 'POST':
        exam_date = request.POST.get('exam_date')
        entered_by = request.POST.get('entered_by', '')
        
        success_count = 0
        
        # Iterate over all possible inputs
        for student in students:
            for subject in subjects:
                # Expected input name format: marks_{student_id}_{subject_id}
                input_name = f'marks_{student.id}_{subject.id}'
                marks_str = request.POST.get(input_name)
                
                if marks_str is not None and marks_str.strip() != '':
                    try:
                        marks_obtained = float(marks_str)
                        
                        # Get max marks: form override → exam-specific override → subject default
                        max_marks_input = f'max_marks_{subject.id}'
                        max_marks_str = request.POST.get(max_marks_input)
                        if max_marks_str and max_marks_str.strip():
                            max_marks = float(max_marks_str)
                        else:
                            max_marks = float(subject.effective_max_marks)
                            
                        # Find matching enrollment
                        enrollment = next((e for e in enrollments if e.student == student), None)
                        
                        # Update or create
                        MarkEntry.objects.update_or_create(
                            student=student,
                            exam_type=exam_type,
                            subject=subject,
                            defaults={
                                'enrollment': enrollment,
                                'marks_obtained': marks_obtained,
                                'max_marks': max_marks,
                                'exam_date': exam_date if exam_date else None,
                                'entered_by': entered_by,
                            }
                        )
                        success_count += 1
                    except ValueError:
                        pass # Ignore invalid inputs
                else:
                    # If field is empty and we are doing subject-wise entry, delete existing mark entry if it exists
                    if selected_subject:
                        MarkEntry.objects.filter(
                            student=student,
                            exam_type=exam_type,
                            subject=subject
                        ).delete()
                        
        messages.success(request, f'Successfully saved {success_count} mark entries.')
        # Redirect back to the same page to show updated data
        redirect_url = f"{reverse('students:mark_entry_step3', args=[exam_type.id])}?grade_id={grade_id}" # Use grade_id
        if division_id:
            redirect_url += f"&division_id={division_id}"
        if selected_subject_id:
            redirect_url += f"&subject_id={selected_subject_id}"
        return redirect(redirect_url)

    # Get existing marks to populate the form
    # Format: existing_marks[student_id][subject_id] = marks_obtained
    existing_marks = {}
    
    mark_entries = MarkEntry.objects.filter(
        exam_type=exam_type,
        student__in=students,
        subject__in=subjects
    )
    
    for entry in mark_entries:
        if entry.student_id not in existing_marks:
            existing_marks[entry.student_id] = {}
        existing_marks[entry.student_id][entry.subject_id] = float(entry.marks_obtained)
        
    context = {
        'exam_type': exam_type,
        'grade_id': grade_id, # Pass grade ID
        'grade_name': grade_obj.name, # Pass grade name
        'division_id': division_id,
        'division_name': division_name,
        'students': students,
        'subjects': subjects,
        'all_class_subjects': all_class_subjects,
        'selected_subject': selected_subject,
        'existing_marks': existing_marks,
    }
    return render(request, 'students/mark_entry_step3.html', context)


@role_required(['admin', 'teacher'])
def mark_bulk_import_template(request, exam_type_id):
    """Download an Excel template for bulk mark entry prefilled with students and subjects"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse
    
    exam_type = get_object_or_404(ExamType, id=exam_type_id)
    grade_id = request.GET.get('grade_id')
    division_id = request.GET.get('division_id')
    subject_id = request.GET.get('subject_id')
    
    if not grade_id:
        messages.error(request, 'Grade is required.')
        return redirect('students:mark_entry_step2', exam_type_id=exam_type.id)
        
    grade_obj = get_object_or_404(Grade, id=grade_id)
    division_obj = None
    division_name = "None"
    division_id_str = "none"
    if division_id and division_id != 'None':
        division_obj = get_object_or_404(Division, id=division_id)
        division_name = division_obj.name
        division_id_str = str(division_obj.id)
        
    # Get students for this class via Enrollment
    active_year = AcademicYear.objects.filter(is_active=True).first()
    enrollments_query = Enrollment.objects.filter(academic_year=active_year, student__is_active=True, grade=grade_obj).select_related('student')
    if exam_type.section:
        enrollments_query = enrollments_query.filter(section=exam_type.section)
        
    if division_obj:
        enrollments_query = enrollments_query.filter(division=division_obj)
    else:
        enrollments_query = enrollments_query.filter(division__isnull=True)
        
    enrollments = list(enrollments_query.order_by('student__last_name', 'student__first_name'))
    
    # Get subjects for this class
    subjects_query = Subject.objects.filter(is_active=True, grade=grade_obj)
    if exam_type.subjects.exists():
        subjects_query = subjects_query.filter(id__in=exam_type.subjects.all())
    if exam_type.subject_type != 'all':
        subjects_query = subjects_query.filter(subject_type=exam_type.subject_type)
        
    if division_obj:
        if exam_type.subject_type == 'hadiya':
            subjects_query = subjects_query.filter(division__isnull=True)
        elif exam_type.subject_type == 'division':
            subjects_query = subjects_query.filter(division=division_obj)
        else:
            subjects_query = subjects_query.filter(
                Q(division=division_obj) | Q(subject_type='hadiya', division__isnull=True)
            )
    else:
        subjects_query = subjects_query.filter(division__isnull=True)
        
    subjects = list(subjects_query.order_by('subject_type', 'name'))
    
    # Optional subject wise filter
    selected_subject = None
    if subject_id and subject_id != 'all':
        try:
            subject_id_int = int(subject_id)
            selected_subject = next((s for s in subjects if s.id == subject_id_int), None)
            if selected_subject:
                subjects = [selected_subject]
        except (ValueError, TypeError):
            pass

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Marks Entry"
    
    # Write metadata
    ws.cell(row=1, column=1, value=f"Exam: {exam_type.name} (ID: {exam_type.id})")
    ws.cell(row=2, column=1, value=f"Grade: {grade_obj.name} (ID: {grade_obj.id})")
    ws.cell(row=3, column=1, value=f"Division: {division_name} (ID: {division_id_str})")
    
    # Style metadata
    meta_font = Font(name="Arial", size=10, italic=True, color="555555")
    for r in range(1, 4):
        ws.cell(row=r, column=1).font = meta_font
        
    # Write headers
    headers = ["Register Number", "Student Name"]
    for sub in subjects:
        headers.append(f"{sub.name} (Max: {sub.max_marks}) (ID: {sub.id})")
        
    # Style headers row
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        
    # Prepopulate student list
    data_font = Font(name="Arial", size=11)
    for row_idx, enrollment in enumerate(enrollments, start=6):
        student = enrollment.student
        ws.cell(row=row_idx, column=1, value=student.student_id).font = data_font
        ws.cell(row=row_idx, column=2, value=student.full_name).font = data_font
        
        # Load existing marks
        for col_idx, sub in enumerate(subjects, start=3):
            # check if mark entry exists
            mark_entry = MarkEntry.objects.filter(student=student, exam_type=exam_type, subject=sub).first()
            if mark_entry:
                ws.cell(row=row_idx, column=col_idx, value=float(mark_entry.marks_obtained)).font = data_font
            else:
                ws.cell(row=row_idx, column=col_idx).font = data_font

    # Adjust widths
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 25
    for col_idx in range(3, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 22
        
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"marks_template_{exam_type.name.replace(' ', '_')}_{grade_obj.name.replace(' ', '_')}"
    if division_obj:
        filename += f"_{division_obj.name.replace(' ', '_')}"
    if selected_subject:
        filename += f"_{selected_subject.name.replace(' ', '_')}"
    filename += ".xlsx"
    
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@role_required(['admin', 'teacher'])
def mark_bulk_import(request, exam_type_id):
    """Import marks from Excel template"""
    import openpyxl
    import re
    from django.core.exceptions import ValidationError
    
    exam_type = get_object_or_404(ExamType, id=exam_type_id)
    grade_id = request.GET.get('grade_id')
    division_id = request.GET.get('division_id')
    subject_id = request.GET.get('subject_id')
    
    if not grade_id:
        messages.error(request, 'Grade is required.')
        return redirect('students:mark_entry_step1')
        
    grade_obj = get_object_or_404(Grade, id=grade_id)
    division_obj = None
    if division_id and division_id != 'None':
        division_obj = get_object_or_404(Division, id=division_id)
        
    if request.method == 'POST':
        if 'excel_file' not in request.FILES:
            messages.error(request, 'Please select an Excel file.')
            redirect_url = f"{reverse('students:mark_entry_step3', args=[exam_type.id])}?grade_id={grade_id}"
            if division_id:
                redirect_url += f"&division_id={division_id}"
            return redirect(redirect_url)
            
        excel_file = request.FILES['excel_file']
        
        try:
            wb = openpyxl.load_workbook(excel_file, data_only=True)
            ws = wb.active
            
            # 1. Parse and Validate Metadata
            exam_meta = str(ws.cell(row=1, column=1).value or '')
            grade_meta = str(ws.cell(row=2, column=1).value or '')
            div_meta = str(ws.cell(row=3, column=1).value or '')
            
            # Extract IDs from metadata
            exam_id_match = re.search(r'\(ID:\s*(\d+)\)', exam_meta)
            grade_id_match = re.search(r'\(ID:\s*(\d+)\)', grade_meta)
            div_id_match = re.search(r'\(ID:\s*([a-zA-Z0-9_]+)\)', div_meta)
            
            if not exam_id_match or not grade_id_match or not div_id_match:
                messages.error(request, 'Invalid template format. Metadata rows 1-3 are missing or corrupted.')
                raise ValidationError('Metadata parsing failed')
                
            parsed_exam_id = int(exam_id_match.group(1))
            parsed_grade_id = int(grade_id_match.group(1))
            parsed_div_id_str = div_id_match.group(1)
            
            if parsed_exam_id != exam_type.id:
                messages.error(request, f'Template exam ID ({parsed_exam_id}) does not match current exam ({exam_type.name}).')
                raise ValidationError('Exam mismatch')
                
            if parsed_grade_id != grade_obj.id:
                messages.error(request, f'Template grade ID ({parsed_grade_id}) does not match selected grade ({grade_obj.name}).')
                raise ValidationError('Grade mismatch')
                
            # Check division
            if division_obj and str(division_obj.id) != parsed_div_id_str:
                messages.error(request, f'Template division ID ({parsed_div_id_str}) does not match selected division ({division_obj.name}).')
                raise ValidationError('Division mismatch')
            elif not division_obj and parsed_div_id_str != 'none':
                messages.error(request, f'Template division mismatch. Expected no division but found ID: {parsed_div_id_str}.')
                raise ValidationError('Division mismatch')
                
            # 2. Parse Headers to find subjects
            headers = []
            col = 1
            while True:
                val = ws.cell(row=5, column=col).value
                if val is None:
                    break
                headers.append((col, str(val).strip()))
                col += 1
                
            if len(headers) < 2 or headers[0][1] != 'Register Number' or headers[1][1] != 'Student Name':
                messages.error(request, 'Template header format mismatch. First columns must be "Register Number" and "Student Name".')
                raise ValidationError('Header mismatch')
                
            subject_cols = []
            for col_idx, header_text in headers[2:]:
                sub_id_match = re.search(r'\(ID:\s*(\d+)\)', header_text)
                if sub_id_match:
                    sub_id = int(sub_id_match.group(1))
                    try:
                        subject = Subject.objects.get(id=sub_id, grade=grade_obj)
                        subject_cols.append((col_idx, subject))
                    except Subject.DoesNotExist:
                        pass # Ignore or warning
                        
            if not subject_cols:
                messages.error(request, 'No valid subjects found in Excel template columns.')
                raise ValidationError('No subjects')
                
            # 3. Process Student Rows
            active_year = AcademicYear.objects.filter(is_active=True).first()
            success_count = 0
            row_idx = 6
            while True:
                student_id = ws.cell(row=row_idx, column=1).value
                if student_id is None:
                    break
                student_id = str(student_id).strip()
                if not student_id:
                    row_idx += 1
                    continue
                    
                # Verify student and enrollment
                try:
                    student = Student.objects.get(student_id=student_id, is_active=True)
                    enrollment = Enrollment.objects.filter(
                        student=student, 
                        academic_year=active_year, 
                        grade=grade_obj
                    ).first()
                    
                    if not enrollment:
                        # Skip or report error
                        row_idx += 1
                        continue
                        
                    # Import marks for this student
                    for col_idx, subject in subject_cols:
                        mark_val = ws.cell(row=row_idx, column=col_idx).value
                        
                        if mark_val is not None and str(mark_val).strip() != '':
                            try:
                                marks_obtained = float(mark_val)
                                max_marks = float(subject.max_marks)
                                
                                MarkEntry.objects.update_or_create(
                                    student=student,
                                    exam_type=exam_type,
                                    subject=subject,
                                    defaults={
                                        'enrollment': enrollment,
                                        'marks_obtained': marks_obtained,
                                        'max_marks': max_marks,
                                        'entered_by': request.user.username,
                                    }
                                )
                                success_count += 1
                            except ValueError:
                                pass # Skip invalid numeric marks
                        else:
                            # If we are doing single subject entry (subject_id query parameter set)
                            # and the mark is empty, delete the existing mark entry
                            if subject_id and subject_id != 'all':
                                MarkEntry.objects.filter(
                                    student=student,
                                    exam_type=exam_type,
                                    subject=subject
                                ).delete()
                                
                except Student.DoesNotExist:
                    pass # Skip invalid students
                    
                row_idx += 1
                
            messages.success(request, f'Excel upload processed successfully. Saved/updated {success_count} mark entries.')
            
        except Exception as e:
            if not messages.get_messages(request):
                messages.error(request, f'Error reading or importing Excel file: {str(e)}')
                
    redirect_url = f"{reverse('students:mark_entry_step3', args=[exam_type.id])}?grade_id={grade_id}"
    if division_id:
        redirect_url += f"&division_id={division_id}"
    if subject_id:
        redirect_url += f"&subject_id={subject_id}"
    return redirect(redirect_url)


from collections import defaultdict
from django.shortcuts import render
from .models import MarkEntry


from django.shortcuts import render
from .models import ExamType, Division

@role_required(['admin', 'teacher'])
def mark_entry_list(request):
    context = {
        'exam_types': ExamType.objects.select_related('section').all(),
        'sections': Section.objects.all().order_by('order', 'name'),
        'divisions': Division.objects.all(),
        'grades': Grade.objects.all().order_by('order', 'name'), # Pass Grade objects
    }
    return render(request, 'students/mark_entry_classwise.html', context)

from collections import defaultdict
from django.http import JsonResponse
from django.template.loader import render_to_string
from .models import MarkEntry


@role_required(['admin', 'teacher'])
def mark_entry_classwise_data(request):
    """
    AJAX view:
    Returns class-wise mark tables with ranking
    """
    mark_entries = MarkEntry.objects.select_related(
        'student',
        'enrollment',
        'enrollment__section',
        'enrollment__division',
        'enrollment__grade', # Select grade
        'exam_type',
        'subject',
    )

    # -------------------------
    # Filters (no reload)
    # -------------------------
    exam_type = request.GET.get('exam_type')
    section = request.GET.get('section')
    grade_id = request.GET.get('grade') # Get grade ID
    division = request.GET.get('division')

    if exam_type:
        mark_entries = mark_entries.filter(exam_type_id=exam_type)
    if section:
        mark_entries = mark_entries.filter(enrollment__section_id=section)
    if grade_id:
        mark_entries = mark_entries.filter(enrollment__grade_id=grade_id) # Filter by grade ID
    if division:
        mark_entries = mark_entries.filter(enrollment__division_id=division)

    # -------------------------
    # Group entries class-wise
    # -------------------------
    grouped = defaultdict(list)

    for entry in mark_entries:
        key = (
            entry.enrollment.section.name if entry.enrollment and entry.enrollment.section else "-",
            entry.enrollment.grade.name if entry.enrollment and entry.enrollment.grade else '-', # Use grade name
            entry.enrollment.division.name if entry.enrollment and entry.enrollment.division else "-",
            entry.exam_type.name,
        )
        grouped[key].append(entry)

    class_tables = []

    # -------------------------
    # Build tables
    # -------------------------
    for (sec, cls, div, exam), entries in grouped.items():

        # Collect subjects (ordered, unique)
        subjects = []
        for e in entries:
            if e.subject.name not in subjects:
                subjects.append(e.subject.name)

        # Collect students
        students = {}

        for e in entries:
            sid = e.student.id
            if sid not in students:
                students[sid] = {
                    'name': e.student.full_name,
                    'marks': [
                        {'marks': '-', 'grade': 'F', 'is_fail': True}
                        for _ in subjects
                    ],
                    'total': 0,
                    'fail_count': len(subjects),  # assume fail until proven pass

                }


        # Fill marks
        for e in entries:
            student = students[e.student.id]
            subject_index = subjects.index(e.subject.name)

            is_fail = str(e.grade_letter).strip().upper() == 'F'

            student['marks'][subject_index] = {
                'marks': float(e.marks_obtained),
                'grade': e.grade_letter,
                'is_fail': is_fail,
            }

            student['total'] += float(e.marks_obtained)

            # reduce fail count if passed
            if not is_fail:
                student['fail_count'] -= 1


        # -------------------------
        # Rank calculation
        # -------------------------
        sorted_students = sorted(
            students.values(),
            key=lambda x: (
                x['fail_count'],   # fewer fails first
                -x['total']        # higher marks next
            )
        )


        prev_key = None
        current_rank = 0

        for index, st in enumerate(sorted_students):
            current_key = (st['fail_count'], st['total'])

            if prev_key is None or current_key != prev_key:
                current_rank = index + 1

            st['rank'] = current_rank
            prev_key = current_key


        # -------------------------
        # Subject-wise averages
        # -------------------------
        subject_totals = [0] * len(subjects)
        student_count = len(sorted_students)

        for st in sorted_students:
            for i, m in enumerate(st['marks']):
                if m['marks'] != '-':
                    subject_totals[i] += m['marks']

        subject_averages = [
            round(total / student_count, 2) if student_count else 0
            for total in subject_totals
        ]

        # -------------------------
        # Final table object
        # -------------------------
        class_tables.append({
            'section': sec,
            'grade': cls,
            'division': div,
            'exam_type': exam,
            'subjects': subjects,
            'students': sorted_students,
            'subject_averages': subject_averages,
        })

    # -------------------------
    # Render partial HTML
    # -------------------------
    html = render_to_string(
        'students/partials/classwise_tables.html',
        {'class_tables': class_tables},
        request=request
    )

    return JsonResponse({'html': html})


@role_required(['admin', 'teacher', 'student'])
def progress_report(request):
    """Generate and view progress reports"""
    student_id = request.GET.get('student_id')
    exam_type_id = request.GET.get('exam_type')
    section_id = request.GET.get('section')
    grade_id = request.GET.get('grade')
    division_id = request.GET.get('division')
    academic_year_name = request.GET.get('academic_year', '')

    # Data isolation for students
    is_student = hasattr(request.user, 'profile') and request.user.profile.role == 'student'
    if is_student:
        if not request.user.profile.student_record:
            messages.error(request, "Student record not found.")
            return redirect('students:home')
        # Force student_id filter to own student_id
        student_id = request.user.profile.student_record.student_id
        
    # Get the active academic year (or most recent if none is active)
    active_year = AcademicYear.objects.filter(is_active=True).first()
    if not active_year:
        active_year = AcademicYear.objects.order_by('-start_date').first()

    if academic_year_name:
        year_obj = AcademicYear.objects.filter(name=academic_year_name).first()
        if year_obj:
            active_year = year_obj

    enrollments_query = Enrollment.objects.filter(academic_year=active_year).select_related('student', 'grade', 'division', 'section')

    if student_id:
        enrollments_query = enrollments_query.filter(student__student_id__icontains=student_id)
    if section_id:
        enrollments_query = enrollments_query.filter(section_id=section_id)
    if grade_id:
        enrollments_query = enrollments_query.filter(grade_id=grade_id) # Filter by grade ID
    if division_id:
        enrollments_query = enrollments_query.filter(division_id=division_id)

    enrollments = enrollments_query

    if request.method == 'POST':
        if is_student:
             messages.error(request, "Students cannot generate progress reports.")
             return redirect('students:progress_report')
        student_ids = request.POST.getlist('students')
        exam_type_id = request.POST.get('exam_type')
        academic_year_name = request.POST.get('academic_year', '') # Changed to academic_year_name
        generated_by = request.user.get_full_name() or request.user.username

        if not student_ids or not exam_type_id:
            messages.error(request, 'Please select students and exam type.')
            return redirect('students:progress_report')

        exam_type = get_object_or_404(ExamType, id=exam_type_id)
        created_count = 0

        for student_id in student_ids:
            student = get_object_or_404(Student, id=student_id)

            # Calculate totals from mark entries
            mark_entries = MarkEntry.objects.filter(
                student=student,
                exam_type=exam_type
            )

            if not mark_entries.exists():
                continue

            total_marks = sum(entry.marks_obtained for entry in mark_entries)
            total_max_marks = sum(entry.max_marks for entry in mark_entries)
            percentage = (total_marks / total_max_marks * 100) if total_max_marks > 0 else 0

            # Calculate grade
            if percentage >= 90:
                grade_letter = 'A+'
            elif percentage >= 80:
                grade_letter = 'A'
            elif percentage >= 70:
                grade_letter = 'B+'
            elif percentage >= 60:
                grade_letter = 'B'
            elif percentage >= 50:
                grade_letter = 'C+'
            elif percentage >= 40:
                grade_letter = 'C'
            else:
                grade_letter = 'F'

            # Find the active enrollment for this student
            enrollment = Enrollment.objects.filter(student=student, academic_year=active_year).first()

            # Create or update progress report
            report, created = ProgressReport.objects.update_or_create(
                student=student,
                exam_type=exam_type,
                enrollment=enrollment,
                defaults={
                    'total_marks_obtained': total_marks,
                    'total_max_marks': total_max_marks,
                    'overall_percentage': percentage,
                    'overall_grade': grade_letter,
                    'generated_by': generated_by,
                }
            )
            if created:
                created_count += 1

        messages.success(request, f'Successfully generated {created_count} progress report(s).')
        return redirect('students:progress_report')

    # Get progress reports
    latest_reports = (
        ProgressReport.objects
        .values('student_id')
        .annotate(latest_time=Max('generated_at'))
    )

    reports = ProgressReport.objects.filter(
        generated_at__in=[r['latest_time'] for r in latest_reports]
    ).select_related(
        'student', 'exam_type', 'enrollment__grade' # Select grade
    )

    if is_student:
        reports = reports.filter(student=request.user.profile.student_record, exam_type__is_published=True)
        student_id = None


    if student_id:
        reports = reports.filter(student__student_id__icontains=student_id)
    if exam_type_id:
        reports = reports.filter(exam_type_id=exam_type_id)
    if grade_id:
        reports = reports.filter(enrollment__grade_id=grade_id) # Filter by grade ID
    if division_id:
        reports = reports.filter(enrollment__division_id=division_id)
    if academic_year_name:
        reports = reports.filter(enrollment__academic_year__name=academic_year_name)

    reports = reports.order_by('-generated_at', 'enrollment__grade__order', 'enrollment__grade__name', 'student__last_name') # Order by grade object

    exam_types = ExamType.objects.all()
    divisions = Division.objects.all()
    sections = Section.objects.all().order_by('order', 'name')
    # Get all unique grades from database
    all_grades = Grade.objects.all().order_by('order', 'name')

    context = {
        'enrollments': enrollments,
        'reports': reports,
        'exam_types': exam_types,
        'divisions': divisions,
        'grades': all_grades,
        'sections': sections,
        'academic_years': AcademicYear.objects.order_by('-start_date'),
        'active_year': active_year,
        'current_filters': {
            'student_id': student_id,
            'exam_type': exam_type_id,
            'section': section_id,
            'grade': grade_id,
            'division': division_id,
            'academic_year': academic_year_name,
        }
    }
    return render(request, 'students/progress_report.html', context)


@role_required(['admin', 'teacher', 'student'])
def progress_report_detail(request, pk):
    """View detailed progress report for a student"""
    report = get_object_or_404(ProgressReport, pk=pk)
    
    # Data isolation for students
    if hasattr(request.user, 'profile') and request.user.profile.role == 'student':
        if not report.exam_type.is_published:
            messages.error(request, "Results for this exam have not been published yet.")
            return redirect('students:home')
        if not request.user.profile.student_record or request.user.profile.student_record != report.student:
            messages.error(request, "You do not have permission to view other students' progress reports.")
            return redirect('students:home')
            
    mark_entries = MarkEntry.objects.filter(
        student=report.student,
        exam_type=report.exam_type
    ).select_related('subject').order_by('subject__subject_type', 'subject__name')

    # Separate Hadiya and division subjects
    hadiya_subjects = [entry for entry in mark_entries if entry.subject.subject_type == 'hadiya']
    division_subjects = [entry for entry in mark_entries if entry.subject.subject_type == 'division']

    context = {
        'report': report,
        'hadiya_subjects': hadiya_subjects,
        'division_subjects': division_subjects,
        'all_subjects': mark_entries,
    }
    return render(request, 'students/progress_report_detail.html', context)

@login_required
def student_profile(request, pk):
    """View a student's full profile (read-only)"""
    student = get_object_or_404(Student, id=pk)
    
    # Ensure students can only view their own profile, unless they are staff/admin
    if hasattr(request.user, 'profile') and request.user.profile.role == 'student':
        if not hasattr(request.user.profile, 'student_record') or request.user.profile.student_record != student:
            messages.error(request, 'You do not have permission to view this profile.')
            return redirect('students:home')
            
    # Get active components
    active_year = AcademicYear.objects.filter(is_active=True).first()
    enrollment = student.enrollments.filter(academic_year=active_year).first()
    
    # If no active enrollment found, just get the most recent one
    if not enrollment and student.enrollments.exists():
        enrollment = student.enrollments.order_by('-academic_year__start_date').first()
        
    # Get today's daily attendance
    today = date.today()
    today_attendance = Attendance.objects.filter(student=student, date=today, attendance_type='daily').first()
    
    # Get progress reports
    progress_reports = ProgressReport.objects.filter(student=student).select_related('exam_type').order_by('-generated_at')
    if hasattr(request.user, 'profile') and request.user.profile.role == 'student':
        progress_reports = progress_reports.filter(exam_type__is_published=True)
        
    context = {
        'student': student,
        'enrollment': enrollment,
        'active_year': active_year,
        'today_attendance': today_attendance,
        'progress_reports': progress_reports,
    }
    return render(request, 'students/student_profile.html', context)

@student_own_data_required
def student_edit(request, pk):
    """Edit an existing student"""
    student = get_object_or_404(Student, id=pk)
    
    # Get all academic years for the dropdown
    academic_years = AcademicYear.objects.all().order_by('-start_date')
    active_year = AcademicYear.objects.filter(is_active=True).first()
    
    # Get current enrollment (prefer the active year, or the most recent one if no active year)
    # But wait, if they are editing a specific enrollment for a specific year, 
    # we should ideally pass that year. For now, we look at active year or latest.
    enrollment = None
    selected_year = active_year
    
    if request.GET.get('year'):
        selected_year = AcademicYear.objects.filter(id=request.GET.get('year')).first()
        
    if selected_year:
        enrollment = Enrollment.objects.filter(student=student, academic_year=selected_year).first()
    elif student.enrollments.exists():
        enrollment = student.enrollments.first()
        selected_year = enrollment.academic_year

    if request.method == 'POST':
        student_id = request.POST.get('student_id')
        first_name = request.POST.get('first_name')
        last_name = request.POST.get('last_name')
        grade_id = request.POST.get('grade')
        division_id = request.POST.get('division') or None
        room_id = request.POST.get('room') or None
        
        section_id = request.POST.get('section') or None
        if grade_id:
            grade_obj = Grade.objects.filter(id=grade_id).first()
            if grade_obj and grade_obj.section:
                section_id = grade_obj.section.id

        student_type = request.POST.get('student_type', 'day_scholar')
        email = request.POST.get('email', '')
        phone = request.POST.get('phone', '')
        address = request.POST.get('address', '')
        bus_stop_id = request.POST.get('bus_stop') if student_type == 'day_scholar' else None
        
        # Get the selected academic year from the form
        form_year_id = request.POST.get('academic_year_id')
        form_year = None
        if form_year_id:
            form_year = AcademicYear.objects.filter(id=form_year_id).first()
        else:
            form_year = active_year

        is_student = hasattr(request.user, 'profile') and request.user.profile.role == 'student'

        if is_student:
            student.email = email
            student.phone = phone
            student.address = address
            student.save()

            if request.user.email != email:
                request.user.email = email
                request.user.save()

            # Optional password update in student edit
            current_password = request.POST.get('current_password', '').strip()
            new_password = request.POST.get('new_password', '').strip()
            confirm_password = request.POST.get('confirm_password', '').strip()

            if current_password or new_password or confirm_password:
                if not current_password:
                    messages.error(request, "Current password is required to change password.")
                    return redirect('students:student_profile', pk=student.id)
                if not request.user.check_password(current_password):
                    messages.error(request, "Current password is incorrect.")
                    return redirect('students:student_profile', pk=student.id)
                if new_password and new_password == confirm_password:
                    request.user.set_password(new_password)
                    request.user.save()
                    profile = request.user.profile
                    profile.initial_password = new_password
                    profile.save()
                    from django.contrib.auth import update_session_auth_hash
                    update_session_auth_hash(request, request.user)
                    messages.success(request, 'Your profile and password have been updated successfully!')
                else:
                    messages.error(request, 'Password confirmation mismatch.')
                    return redirect('students:student_profile', pk=student.id)
            else:
                messages.success(request, 'Your profile details were updated successfully!')
            
            return redirect('students:student_profile', pk=student.id)
        else:
            # Check duplicate student ID, but allow same ID for this student
            if Student.objects.exclude(id=student.id).filter(student_id=student_id).exists():
                messages.error(request, f'Student ID {student_id} already exists.')
            else:
                student.student_id = student_id
                student.first_name = first_name
                student.last_name = last_name
                student.student_type = student_type
                student.email = email
                student.phone = phone
                student.address = address
                student.bus_stop_id = bus_stop_id
                student.save()
                
                siblings_ids = request.POST.getlist('siblings')
                student.siblings.set(siblings_ids)
                
                if form_year:
                    # Find if an enrollment already exists for this specific student and year
                    year_enrollment = Enrollment.objects.filter(student=student, academic_year=form_year).first()
                    if year_enrollment:
                        year_enrollment.grade_id = grade_id
                        year_enrollment.section_id = section_id
                        year_enrollment.division_id = division_id
                        year_enrollment.room_id = room_id
                        year_enrollment.save()
                    else:
                        Enrollment.objects.create(
                            student=student,
                            academic_year=form_year,
                            grade_id=grade_id,
                            section_id=section_id,
                            division_id=division_id,
                            room_id=room_id
                        )

                messages.success(request, f'Student {student.full_name} updated successfully for {form_year.name if form_year else "the selected year"}!')
                return redirect('students:student_list')

    divisions = Division.objects.all()
    rooms = Room.objects.all()
    sections = Section.objects.all().order_by('order', 'name')
    grades = Grade.objects.all().order_by('order', 'name')
    all_students = Student.objects.filter(is_active=True).exclude(id=student.id).order_by('first_name', 'last_name')
    from fees.models import BusStop
    bus_stops = BusStop.objects.all().order_by('stop_name')

    context = {
        'student': student,
        'divisions': divisions,
        'rooms': rooms,
        'sections': sections,
        'grades': grades,
        'all_students': all_students,
        'academic_years': academic_years,
        'selected_year': selected_year,
        'active_year': active_year,
        'enrollment': enrollment,
        'bus_stops': bus_stops,
    }
    return render(request, 'students/student_edit.html', context)

@role_required(['admin', 'ntstaff'])
def create_student_user(request, pk):
    """Create a linked User account for an existing student"""
    if request.method == 'POST':
        student = get_object_or_404(Student, id=pk)
        
        # Check if already linked
        if hasattr(student, 'user_profile') and student.user_profile:
            messages.error(request, f'Student {student.full_name} already has a linked user account.')
            return redirect('students:student_edit', pk=pk)
            
        username = request.POST.get('username', '').strip()
        password = request.POST.get('password', '')
        
        if not username or not password:
            messages.error(request, 'Username and password are required.')
            return redirect('students:student_edit', pk=pk)
            
        # Check if username exists
        if User.objects.filter(username=username).exists():
            messages.error(request, f'Username "{username}" is already taken.')
            return redirect('students:student_edit', pk=pk)
            
        try:
            # Create the auth User
            user = User.objects.create_user(
                username=username,
                password=password,
                first_name=student.first_name,
                last_name=student.last_name,
                email=student.email or ''
            )
            
            # The signal automatically created a UserProfile for this user with role='ntstaff'.
            # We need to fetch it and update it to 'student' and link the student record.
            profile = user.profile
            profile.role = 'student'
            profile.student_record = student
            profile.save()
            
            messages.success(request, f'Login account created successfully for {student.full_name}!')
            
        except Exception as e:
            messages.error(request, f'Error creating user account: {str(e)}')
            
    return redirect('students:student_edit', pk=pk)


@role_required(['admin'])
def student_delete(request, pk):
    student = get_object_or_404(Student, pk=pk)
    student.delete()
    messages.success(request, "Student deleted successfully.")
    return redirect('students:student_list')

from django.shortcuts import render
from django.db.models import Count, Q, Avg
from django.utils import timezone
from datetime import datetime, timedelta, date
from collections import defaultdict
from calendar import monthrange


@role_required(['admin', 'teacher'])
def attendance_analytics(request):
    """Enhanced attendance analytics with multiple time ranges"""

    # Get filter parameters
    report_type = request.GET.get('report_type', 'daily')  # daily, monthly, quarterly, yearly, custom
    selected_date = request.GET.get('date', date.today().isoformat())
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    section_id = request.GET.get('section')
    grade = request.GET.get('grade')
    division_id = request.GET.get('division')

    # Parse selected date
    try:
        current_date = datetime.strptime(selected_date, '%Y-%m-%d').date()
    except:
        current_date = date.today()

    # Calculate date ranges based on report type
    if report_type == 'daily':
        start_date = current_date
        end_date = current_date
    elif report_type == 'monthly':
        start_date = current_date.replace(day=1)
        _, last_day = monthrange(current_date.year, current_date.month)
        end_date = current_date.replace(day=last_day)
    elif report_type == 'quarterly':
        # Get current quarter
        quarter = (current_date.month - 1) // 3 + 1
        start_month = (quarter - 1) * 3 + 1
        start_date = date(current_date.year, start_month, 1)
        end_month = start_month + 2
        _, last_day = monthrange(current_date.year, end_month)
        end_date = date(current_date.year, end_month, last_day)
    elif report_type == 'yearly':
        start_date = date(current_date.year, 1, 1)
        end_date = date(current_date.year, 12, 31)
    elif report_type == 'custom' and date_from and date_to:
        start_date = datetime.strptime(date_from, '%Y-%m-%d').date()
        end_date = datetime.strptime(date_to, '%Y-%m-%d').date()
    else:
        start_date = current_date
        end_date = current_date

    # Base queryset
    attendances = Attendance.objects.filter(
        date__gte=start_date,
        date__lte=end_date,
        enrollment__academic_year__is_active=True
    ).select_related('student', 'enrollment', 'enrollment__division', 'enrollment__room')

    # Apply filters
    if section_id:
        attendances = attendances.filter(enrollment__section_id=section_id)
    if grade:
        attendances = attendances.filter(enrollment__grade=grade)
    if division_id:
        attendances = attendances.filter(enrollment__division_id=division_id)

    # Exclude holiday absences from statistics calculations
    h_dates = get_holiday_dates(start_date, end_date)
    stats_attendances = attendances.exclude(date__in=h_dates, status='absent')

    # Calculate overall statistics
    total_records = stats_attendances.count()
    present_count = stats_attendances.filter(status='present').count()
    absent_count = stats_attendances.filter(status='absent').count()
    late_count = stats_attendances.filter(status='late').count()
    excused_count = stats_attendances.filter(status='excused').count()

    overall_percentage = round((present_count / total_records * 100), 2) if total_records > 0 else 0
    present_percentage = round((present_count / total_records * 100), 2) if total_records > 0 else 0
    absent_percentage = round((absent_count / total_records * 100), 2) if total_records > 0 else 0
    late_percentage = round((late_count / total_records * 100), 2) if total_records > 0 else 0
    excused_percentage = round((excused_count / total_records * 100), 2) if total_records > 0 else 0

    # Division-wise breakdown
    division_stats = {}

    # Get distinct combinations of section, grade and division
    grade_div_combinations = stats_attendances.values_list(
        'enrollment__section__id',
        'enrollment__section__name',
        'enrollment__grade__id',
        'enrollment__grade__name',
        'enrollment__division__id',
        'enrollment__division__name'
    ).distinct()

    for sec_id, sec_name, grade_id_val, grade_name_val, div_id, div_name in grade_div_combinations:
        # Filter attendances for this specific grade + division combination
        div_attendances = stats_attendances.filter(
            enrollment__section_id=sec_id,
            enrollment__grade_id=grade_id_val,
            enrollment__division_id=div_id
        )

        div_total = div_attendances.count()
        div_present = div_attendances.filter(status='present').count()
        div_absent = div_attendances.filter(status='absent').count()
        div_late = div_attendances.filter(status='late').count()
        div_excused = div_attendances.filter(status='excused').count()

        div_percentage = round((div_present / div_total * 100), 2) if div_total > 0 else 0

        # Create a unique key for section + grade + division
        sec_name_display = sec_name or 'No Section'
        key = f"{sec_name_display} - {grade_name_val or 'N/A'} - {div_name or 'No Division'}"

        division_stats[key] = {
            'id': div_id,
            'section': sec_name_display,
            'grade_id': grade_id_val,
            'grade': grade_name_val,
            'division': div_name,
            'total': div_total,
            'present': div_present,
            'absent': div_absent,
            'late': div_late,
            'excused': div_excused,
            'percentage': div_percentage,
            'status': 'good' if div_percentage >= 85 else 'warning' if div_percentage >= 75 else 'critical'
        }

    # Sort division_stats by section, grade and division name
    division_stats = dict(sorted(division_stats.items(), key=lambda x: (x[1]['section'] or '', x[1]['grade'] or '', x[1]['division'] or '')))

    # Student-wise statistics (for detailed reports)
    student_stats = []
    if report_type in ['monthly', 'quarterly', 'yearly', 'custom']:
        enrollments = Enrollment.objects.filter(academic_year__is_active=True, student__is_active=True).select_related('student', 'division', 'grade')
        if grade:
            enrollments = enrollments.filter(grade_id=grade)
        if division_id:
            enrollments = enrollments.filter(division_id=division_id)

        for enrollment in enrollments[:100]:  # Limit to avoid performance issues
            student_attendances = stats_attendances.filter(student=enrollment.student)
            total = student_attendances.count()
            present = student_attendances.filter(status='present').count()
            absent = student_attendances.filter(status='absent').count()
            late = student_attendances.filter(status='late').count()

            if total > 0:
                percentage = round((present / total * 100), 2)
                student_stats.append({
                    'student': enrollment.student,
                    'enrollment': enrollment,
                    'total': total,
                    'present': present,
                    'absent': absent,
                    'late': late,
                    'percentage': percentage,
                    'status': 'good' if percentage >= 85 else 'warning' if percentage >= 75 else 'critical'
                })

        # Sort by percentage (lowest first for attention)
        student_stats.sort(key=lambda x: x['percentage'])

    # Daily trend (last 7 days for daily/monthly views)
    daily_trend = []
    if report_type in ['daily', 'monthly']:
        for i in range(6, -1, -1):
            trend_date = current_date - timedelta(days=i)
            day_attendances = Attendance.objects.filter(date=trend_date, enrollment__academic_year__is_active=True)
            if section_id:
                day_attendances = day_attendances.filter(enrollment__section_id=section_id)
            if grade:
                day_attendances = day_attendances.filter(enrollment__grade_id=grade)
            if division_id:
                day_attendances = day_attendances.filter(enrollment__division_id=division_id)

            h_dates_trend = get_holiday_dates(trend_date, trend_date)
            if trend_date in h_dates_trend:
                day_attendances = day_attendances.exclude(status='absent')

            day_total = day_attendances.count()
            day_present = day_attendances.filter(status='present').count()
            day_percentage = round((day_present / day_total * 100), 2) if day_total > 0 else 0

            daily_trend.append({
                'date': trend_date,
                'percentage': day_percentage,
                'total': day_total
            })

    # Get available grades and divisions for filters
    grades = Grade.objects.all().order_by('order', 'name')
    divisions = Division.objects.all()

    # Handle CSV export
    if request.GET.get('export') == 'csv':
        import csv
        from django.http import HttpResponse

        response = HttpResponse(content_type='text/csv')
        filename = f'attendance_report_{report_type}_{start_date}_to_{end_date}.csv'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        writer = csv.writer(response)

        # Write header
        writer.writerow(['Attendance Report'])
        writer.writerow([f'Period: {start_date} to {end_date}'])
        writer.writerow([f'Report Type: {report_type}'])
        writer.writerow([])

        # Overall statistics
        writer.writerow(['Overall Statistics'])
        writer.writerow(['Metric', 'Count', 'Percentage'])
        writer.writerow(['Total Records', total_records, ''])
        writer.writerow(['Present', present_count, f'{overall_percentage}%'])
        writer.writerow(['Absent', absent_count, f'{round(absent_count/total_records*100, 2) if total_records > 0 else 0}%'])
        writer.writerow(['Late', late_count, f'{round(late_count/total_records*100, 2) if total_records > 0 else 0}%'])
        writer.writerow(['Excused', excused_count, f'{round(excused_count/total_records*100, 2) if total_records > 0 else 0}%'])
        writer.writerow([])

        # Division-wise statistics
        if division_stats:
            writer.writerow(['Division-wise Statistics'])
            writer.writerow(['Class Details', 'Total', 'Present', 'Absent', 'Late', 'Excused', 'Percentage', 'Status'])
            for key, stats in division_stats.items():
                writer.writerow([
                    key,
                    stats['total'],
                    stats['present'],
                    stats['absent'],
                    stats['late'],
                    stats['excused'],
                    f"{stats['percentage']}%",
                    stats['status']
                ])
            writer.writerow([])

        # Student-wise statistics
        if student_stats:
            writer.writerow(['Student-wise Statistics'])
            writer.writerow(['Student ID', 'Name', 'Section', 'Grade', 'Division', 'Total', 'Present', 'Absent', 'Late', 'Percentage', 'Status'])
            for stat in student_stats:
                writer.writerow([
                    stat['student'].student_id,
                    stat['student'].full_name,
                    stat['enrollment'].section.name if stat['enrollment'].section else '-',
                    stat['enrollment'].grade,
                    stat['enrollment'].division.name if stat['enrollment'].division else '-',
                    stat['total'],
                    stat['present'],
                    stat['absent'],
                    stat['late'],
                    f"{stat['percentage']}%",
                    stat['status']
                ])

        return response

    context = {
        'report_type': report_type,
        'selected_date': selected_date,
        'date_from': date_from,
        'date_to': date_to,
        'selected_grade': grade,
        'selected_division': division_id,
        'selected_section': section_id,
        'grades': grades,
        'divisions': divisions,
        'sections': Section.objects.all().order_by('order', 'name'),
        'total_records': total_records,
        'present_count': present_count,
        'absent_count': absent_count,
        'late_count': late_count,
        'excused_count': excused_count,
        'overall_percentage': overall_percentage,
        'present_percentage': present_percentage,
        'absent_percentage': absent_percentage,
        'late_percentage': late_percentage,
        'excused_percentage': excused_percentage,
        'division_stats': division_stats,
        'student_stats': student_stats,
        'daily_trend': daily_trend,
        'start_date': start_date,
        'end_date': end_date,
    }

    return render(request, 'students/attendance_analytics.html', context)

# Detect students absent 3+ days in a row
def get_consecutive_absences(student_id, days=3):
    recent = Attendance.objects.filter(
        student_id=student_id,
        date__gte=date.today() - timedelta(days=days)
    ).order_by('-date')

    return all(a.status == 'absent' for a in recent[:days])

from collections import defaultdict
from django.shortcuts import render
from .models import MarkEntry, ExamType

@role_required(['admin', 'teacher'])
def performance_analysis(request):
    exam_type_id = request.GET.get('exam_type')

    exam_types = ExamType.objects.all()
    toppers_by_class = []

    selected_exam_obj = None
    if exam_type_id:
        selected_exam_obj = ExamType.objects.filter(id=exam_type_id).first()


    if exam_type_id:
        entries = MarkEntry.objects.select_related(
            'student',
            'enrollment',
            'enrollment__section',
            'enrollment__division',
            'exam_type',
            'subject',
        ).filter(exam_type_id=exam_type_id)

        grouped = defaultdict(list)

        # Group by section + class + division
        for e in entries:
            key = (
                e.enrollment.section.name if e.enrollment and e.enrollment.section else "-",
                e.enrollment.grade if e.enrollment else "-",
                e.enrollment.division.name if e.enrollment and e.enrollment.division else "-"
            )
            grouped[key].append(e)

        for (section, grade, division), marks in grouped.items():
            students = {}
            subjects_per_student = defaultdict(set)

            # -------------------------
            # Build student-wise data
            # -------------------------
            for m in marks:
                sid = m.student.id

                if sid not in students:
                    students[sid] = {
                        'name': m.student.full_name,
                        'total': 0,
                        'fail_count': 0,
                    }

                students[sid]['total'] += float(m.marks_obtained)
                subjects_per_student[sid].add(m.subject_id)

                # FAIL if grade is F
                is_fail = str(m.grade_letter).strip().upper() == 'F'
                if is_fail:
                    students[sid]['fail_count'] += 1

            # -------------------------
            # ABSENT = FAIL
            # -------------------------
            # find total subjects in this class
            all_subjects = set(m.subject_id for m in marks)

            for sid, st in students.items():
                attended_subjects = subjects_per_student.get(sid, set())
                absent_count = len(all_subjects - attended_subjects)

                st['fail_count'] += absent_count  # absent = fail

            # -------------------------
            # FILTER: ONLY FULL PASS STUDENTS
            # -------------------------
            passed_students = [
                st for st in students.values()
                if st['fail_count'] == 0
            ]

            if not passed_students:
                continue  # no toppers for this class

            # -------------------------
            # SORT & RANK
            # -------------------------
            ranked = sorted(
                passed_students,
                key=lambda x: -x['total']
            )

            for idx, st in enumerate(ranked):
                st['rank'] = idx + 1

            toppers_by_class.append({
                'section': section,
                'grade': grade,
                'division': division,
                'toppers': ranked[:3]  # 🥇🥈🥉 ONLY PASSED
            })

    context = {
        'exam_types': exam_types,
        'selected_exam': exam_type_id,
        'selected_exam_obj': selected_exam_obj,
        'toppers_by_class': toppers_by_class
    }
    return render(request, 'students/performance_analysis.html', context)

from django.http import HttpResponse
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib.units import cm
from reportlab.platypus import Table, TableStyle
from reportlab.lib import colors
from django.db.models import Max
from django.db.models import IntegerField
from django.db.models.functions import Cast



@role_required(['admin', 'teacher'])
def bulk_progress_report_pdf(request):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="progress_reports.pdf"'

    c = canvas.Canvas(response, pagesize=A4)
    width, height = A4

    latest_reports = (
        ProgressReport.objects
        .values('student_id')
        .annotate(latest_time=Max('generated_at'))
    )

    reports = ProgressReport.objects.filter(
        generated_at__in=[r['latest_time'] for r in latest_reports]
    ).select_related(
        'student', 'exam_type', 'enrollment', 'enrollment__division'
    ).annotate(
        student_id_int=Cast('student__student_id', IntegerField())
    ).order_by(
        'student_id_int'
    )


    for report in reports:
        y = height - 2 * cm

        # ================= INSTITUTION HEADER =================
        c.setFont("Helvetica-Bold", 14)
        c.drawCentredString(
            width / 2,
            y,
            "MARKAZ HADIYA WOMEN'S COLLEGE, THAZHAPRA"
        )
        y -= 0.7 * cm

        c.setFont("Helvetica-Bold", 12)
        c.drawCentredString(
            width / 2,
            y,
            "HIGHER SECONDARY HALF YEARLY EXAMINATION"
        )
        y -= 1 * cm

        # Divider line
        c.line(2 * cm, y, width - 2 * cm, y)
        y -= 1 * cm

        # ================= REPORT TITLE =================
        c.setFont("Helvetica-Bold", 16)
        c.drawCentredString(width / 2, y, "PROGRESS REPORT")
        y -= 1.2 * cm

        c.setFont("Helvetica", 11)
        c.drawString(2 * cm, y, f"Student ID : {report.student.student_id}")
        y -= 0.6 * cm
        c.drawString(2 * cm, y, f"Name       : {report.student.full_name}")
        y -= 0.6 * cm
        c.drawString(2 * cm, y, f"Grade      : {report.enrollment.grade.name if report.enrollment and report.enrollment.grade else '-'}")
        y -= 0.6 * cm
        c.drawString(2 * cm, y, f"Division   : {report.enrollment.division.name if report.enrollment and report.enrollment.division else '-'}")
        y -= 0.6 * cm
        c.drawString(2 * cm, y, f"Exam       : {report.exam_type.name}")
        y -= 0.6 * cm
        c.drawString(2 * cm, y, f"Academic Year : {report.enrollment.academic_year.name if report.enrollment else '-'}")
        y -= 1 * cm

        # ================= OVERALL PERFORMANCE =================
        c.setFont("Helvetica-Bold", 12)
        c.drawString(2 * cm, y, "Overall Performance")
        y -= 0.4 * cm

        overall_table = Table(
            [
                ["Total Marks", "Percentage", "Grade", "Rank"],
                [
                    f"{report.total_marks_obtained}/{report.total_max_marks}",
                    f"{report.overall_percentage:.2f}%",
                    report.overall_grade,
                    report.rank or "-"
                ]
            ],
            colWidths=[4 * cm] * 4
        )

        overall_table.setStyle(TableStyle([
            ('GRID', (0,0), (-1,-1), 1, colors.black),
            ('BACKGROUND', (0,0), (-1,0), colors.lightgrey),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('FONT', (0,0), (-1,0), 'Helvetica-Bold'),
        ]))

        overall_table.wrapOn(c, width, height)
        overall_table.drawOn(c, 2 * cm, y - 2 * cm)
        y -= 3 * cm

        # ================= SUBJECT DATA (SAME AS DETAIL VIEW) =================
        mark_entries = MarkEntry.objects.filter(
            student=report.student,
            exam_type=report.exam_type
        ).select_related('subject').order_by(
            'subject__subject_type', 'subject__name'
        )

        hadiya_subjects = [
            e for e in mark_entries if e.subject.subject_type == 'hadiya'
        ]
        division_subjects = [
            e for e in mark_entries if e.subject.subject_type == 'division'
        ]

        def draw_subject_table(title, entries, y_pos):
            if not entries:
                return y_pos

            c.setFont("Helvetica-Bold", 12)
            c.drawString(2 * cm, y_pos, title)
            y_pos -= 0.4 * cm

            data = [["Subject", "Marks", "Max", "Percentage", "Grade"]]

            for e in entries:
                data.append([
                    e.subject.name,
                    e.marks_obtained,
                    e.max_marks,
                    f"{e.percentage:.2f}%",
                    e.grade_letter
                ])

            table = Table(data, colWidths=[6*cm, 3*cm, 3*cm, 3*cm, 2*cm])
            table.setStyle(TableStyle([
                ('GRID', (0,0), (-1,-1), 1, colors.black),
                ('BACKGROUND', (0,0), (-1,0), colors.lightgrey),
                ('ALIGN', (1,1), (-1,-1), 'CENTER'),
                ('FONT', (0,0), (-1,0), 'Helvetica-Bold'),
            ]))

            table_height = len(data) * 0.6 * cm
            table.wrapOn(c, width, height)
            table.drawOn(c, 2 * cm, y_pos - table_height)

            return y_pos - table_height - 0.8 * cm

        y = draw_subject_table("Hadiya (Islamic) Subjects", hadiya_subjects, y)
        y = draw_subject_table("Division Specific Subjects", division_subjects, y)
        # ================= PARENT SIGNATURE =================
        y -= 1.5 * cm

        c.setFont("Helvetica-Bold", 11)
        c.drawString(2 * cm, y, "Parent / Guardian Signature:")
        c.drawString(width / 2 + 1 * cm, y, "Date:")

        y -= 0.8 * cm

        # Signature lines
        c.line(2 * cm, y, width / 2 - 1 * cm, y)
        c.line(width / 2 + 1 * cm, y, width - 2 * cm, y)


        # ================= NEW PAGE =================
        c.showPage()

    c.save()
    return response

from datetime import date
import calendar
from django.shortcuts import render
from students.models import Student
from .models import Attendance
from students.models import Holiday



@role_required(['admin', 'teacher'])
def attendance_update_tracking(request):
    """
    Monthly Attendance Update Tracking
    (Classroom-wise audit page)
    """

    today = date.today()
    month = int(request.GET.get('month', today.month))
    year = int(request.GET.get('year', today.year))

    days_in_month = calendar.monthrange(year, month)[1]
    month_days = list(range(1, days_in_month + 1))

    active_year = AcademicYear.objects.filter(is_active=True).first()
    
    class_divisions = (
        Enrollment.objects
        .filter(academic_year=active_year)
        .exclude(grade__isnull=True)
        .select_related('division', 'section')
        .values('section__id', 'section__name', 'grade', 'division__id', 'division__name')
        .distinct()
        .order_by('section__name', 'grade', 'division__name')
    )

    attendance_matrix = []

    for cd in class_divisions:
        row = {
            'section': cd['section__name'],
            'grade': cd['grade'],
            'division': cd['division__name'],
            'days': []
        }

        for d in month_days:
            current_date = date(year, month, d)
            weekday = current_date.weekday()
            grade_val = cd['grade']
            manual_holiday = Holiday.objects.filter(date=current_date).filter(
                models.Q(grades__isnull=True) | models.Q(grades=grade_val) | models.Q(grades__name=grade_val)
            ).first()

            is_manual_holiday = manual_holiday is not None
            holiday_title = manual_holiday.title if manual_holiday else ''

            # Holiday logic
            is_sunday = weekday == 6
            is_second_saturday = (
                weekday == 5 and
                ((d + date(year, month, 1).weekday()) // 7 + 1) == 2
            )

            is_holiday = is_sunday or is_second_saturday or is_manual_holiday

            recorded = False
            if not is_holiday:
                recorded = Attendance.objects.filter(
                    enrollment__section_id=cd['section__id'],
                    enrollment__grade=cd['grade'],
                    enrollment__division_id=cd['division__id'],
                    enrollment__academic_year=active_year,
                    date=current_date
                ).exists()

            row['days'].append({
                'day': d,
                'recorded': recorded,
                'is_holiday': is_holiday,
                'holiday_type': (
                    holiday_title if is_manual_holiday
                    else 'Sunday' if is_sunday
                    else 'Second Saturday' if is_second_saturday
                    else ''
                )
            })


        attendance_matrix.append(row)

    MONTH_CHOICES = [
        (1, 'January'), (2, 'February'), (3, 'March'),
        (4, 'April'), (5, 'May'), (6, 'June'),
        (7, 'July'), (8, 'August'), (9, 'September'),
        (10, 'October'), (11, 'November'), (12, 'December'),
    ]

    CURRENT_YEAR = date.today().year
    YEAR_CHOICES = list(range(CURRENT_YEAR - 3, CURRENT_YEAR + 2))


    context = {
        'attendance_matrix': attendance_matrix,
        'month_days': month_days,
        'selected_month': month,
        'selected_year': year,
        'month_choices': MONTH_CHOICES,
        'year_choices': YEAR_CHOICES,
    }

    return render(request, 'students/attendance_update_tracking.html', context)
from django.contrib.auth.decorators import login_required


@role_required(['admin', 'teacher'])
def attendance_class_detail(request, grade_id, division_id):
    """
    Shows a list of students for a specific grade and division,
    along with their total attendance count and percentage.
    """
    try:
        if division_id == 0:
            division = None
        else:
            division = Division.objects.get(id=division_id)
    except Division.DoesNotExist:
        messages.error(request, "Division not found.")
        return redirect('students:attendance_list')

    active_year = AcademicYear.objects.filter(is_active=True).first()
    grade = get_object_or_404(Grade, id=grade_id)
    section_id = request.GET.get('section')
    
    if division:
        enrollments = Enrollment.objects.filter(grade_id=grade_id, division=division, academic_year=active_year, student__is_active=True).select_related('student', 'section').order_by('student__first_name', 'student__last_name')
    else:
        enrollments = Enrollment.objects.filter(grade_id=grade_id, division__isnull=True, academic_year=active_year, student__is_active=True).select_related('student', 'section').order_by('student__first_name', 'student__last_name')
        
    if section_id:
        enrollments = enrollments.filter(section_id=section_id)
        section = Section.objects.get(id=section_id)
    else:
        enrollments = enrollments.filter(section__isnull=True)
        section = None
    
    # Use grade session_start_date to limit attendance to class start date
    grade_session_start = grade.session_start_date if grade and grade.session_start_date else (active_year.start_date if active_year else None)

    # Calculate attendance stats for each student
    student_stats = []
    
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')

    # If no manual date_from, use session start date
    effective_date_from = date_from or (grade_session_start.isoformat() if grade_session_start else None)
    
    # Pre-fetch holiday dates in range to optimize performance
    class_attendances = Attendance.objects.filter(enrollment__grade_id=grade_id)
    if division:
        class_attendances = class_attendances.filter(enrollment__division=division)
    else:
        class_attendances = class_attendances.filter(enrollment__division__isnull=True)
        
    if effective_date_from:
        class_attendances = class_attendances.filter(date__gte=effective_date_from)
    if date_to:
        class_attendances = class_attendances.filter(date__lte=date_to)
        
    dates = list(class_attendances.values_list('date', flat=True))
    h_dates = get_holiday_dates(min(dates), max(dates)) if dates else set()
    
    today = timezone.now().date()
    today_attendances = Attendance.objects.filter(enrollment__in=enrollments, date=today)
    today_status_map = {att.enrollment_id: att.status for att in today_attendances}
    
    for enrollment in enrollments:
        attendances = Attendance.objects.filter(enrollment=enrollment)
        
        if effective_date_from:
            attendances = attendances.filter(date__gte=effective_date_from)
        if date_to:
            attendances = attendances.filter(date__lte=date_to)
            
        stats_attendances = attendances.exclude(date__in=h_dates, status='absent')
        
        total_days = stats_attendances.count()
        present_count = stats_attendances.filter(status='present').count()
        late_count = stats_attendances.filter(status='late').count()
        excused_count = stats_attendances.filter(status='excused').count()
        absent_count = stats_attendances.filter(status='absent').count()
        
        # Calculate percentage (Present + Late + Excused are considered 'attended')
        attended_count = present_count + late_count + excused_count
        
        percentage = 0
        if total_days > 0:
            percentage = round((attended_count / total_days) * 100, 2)
            
        today_status = today_status_map.get(enrollment.id, 'not_marked')
            
        student_stats.append({
            'student': enrollment.student,
            'enrollment': enrollment,
            'total_days': total_days,
            'present_count': present_count,
            'late_count': late_count,
            'excused_count': excused_count,
            'absent_count': absent_count,
            'percentage': percentage,
            'today_status': today_status
        })

    context = {
        'grade': grade,
        'division': division,
        'student_stats': student_stats,
        'section': section,
        'active_year': active_year,
        'session_start': grade_session_start,
        'current_filters': {
            'date_from': date_from or '',
            'date_to': date_to or '',
        }
    }
    return render(request, 'students/attendance_class_detail.html', context)


@role_required(['admin', 'teacher', 'student'])
def attendance_student_detail(request, student_id):
    """
    Shows detailed attendance history for a single student.
    Includes academic year switcher, monthly breakdown and date-specific checker.
    """
    # Data isolation for students
    if hasattr(request.user, 'profile') and request.user.profile.role == 'student':
        if not request.user.profile.student_record or request.user.profile.student_record.id != int(student_id):
            messages.error(request, "You do not have permission to view other students' attendance details.")
            return redirect('students:home')
            
    try:
        student = Student.objects.get(id=student_id)
    except Student.DoesNotExist:
        messages.error(request, "Student not found.")
        return redirect('students:attendance_list')

    # All academic years this student was enrolled in, ordered newest first
    all_enrollments = student.enrollments.select_related('academic_year', 'grade', 'division').order_by('-academic_year__start_date')
    all_years = [(e.academic_year, e) for e in all_enrollments]

    # Determine selected year from GET param or default to active year
    selected_year_id = request.GET.get('selected_year')
    active_year = AcademicYear.objects.filter(is_active=True).first()

    if selected_year_id:
        try:
            selected_year_id = int(selected_year_id)
            enrollment = next((e for (ay, e) in all_years if ay.id == selected_year_id), None)
            selected_year = enrollment.academic_year if enrollment else active_year
        except (ValueError, TypeError):
            enrollment = Enrollment.objects.filter(student=student, academic_year=active_year).first()
            selected_year = active_year
    else:
        selected_year = active_year
        enrollment = Enrollment.objects.filter(student=student, academic_year=active_year).first()

    # Use grade's session_start_date if set (limits attendance display to class start)
    grade_obj = enrollment.grade if enrollment else None
    session_start = None
    if grade_obj and grade_obj.session_start_date:
        session_start = grade_obj.session_start_date
    elif selected_year and selected_year.start_date:
        session_start = selected_year.start_date

    # Filter attendance to the selected year's date range
    attendances = Attendance.objects.filter(student=student).order_by('-date')
    if session_start:
        attendances = attendances.filter(date__gte=session_start)
    if selected_year and selected_year.end_date:
        attendances = attendances.filter(date__lte=selected_year.end_date)
    
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    check_date = request.GET.get('check_date')
    
    # For date checker
    check_result = None
    if check_date:
        check_record = Attendance.objects.filter(student=student, date=check_date).first()
        if check_record:
            check_result = {
                'date': check_date,
                'status': check_record.get_status_display(),
                'status_class': check_record.status,
                'marked_by': check_record.marked_by,
                'remarks': check_record.remarks
            }
        else:
             check_result = {
                'date': check_date,
                'status': 'No Record Found',
                'status_class': 'unknown',
                'marked_by': '-',
                'remarks': '-'
             }

    if date_from:
        attendances = attendances.filter(date__gte=date_from)
    if date_to:
        attendances = attendances.filter(date__lte=date_to)

    dates = list(attendances.values_list('date', flat=True))
    h_dates = get_holiday_dates(min(dates), max(dates)) if dates else set()
    stats_attendances = attendances.exclude(date__in=h_dates, status='absent')

    total_days = stats_attendances.count()
    present_count = stats_attendances.filter(status='present').count()
    late_count = stats_attendances.filter(status='late').count()
    excused_count = stats_attendances.filter(status='excused').count()
    absent_count = stats_attendances.filter(status='absent').count()
    
    attended_count = present_count + late_count + excused_count
    percentage = 0
    if total_days > 0:
        percentage = round((attended_count / total_days) * 100, 2)

    context = {
        'student': student,
        'enrollment': enrollment,
        'all_years': all_years,
        'selected_year': selected_year,
        'attendances': attendances,
        'total_days': total_days,
        'present_count': present_count,
        'late_count': late_count,
        'excused_count': excused_count,
        'absent_count': absent_count,
        'percentage': percentage,
        'check_result': check_result,
        'session_start': session_start,
        'current_filters': {
            'date_from': date_from,
            'date_to': date_to,
            'check_date': check_date
        }
    }
    return render(request, 'students/attendance_student_detail.html', context)


# ==========================================
# SECTION MANAGEMENT
# ==========================================

from .models import Section
from .forms import SectionForm

@role_required(['admin'])
def section_list(request):
    """List all sections"""
    sections = Section.objects.all().order_by('order', 'name')
    context = {'sections': sections}
    return render(request, 'students/section_list.html', context)


@role_required(['admin', 'ntstaff'])
def hostel_student_list_view(request):
    """List all students registered to the hostel"""
    if hasattr(request.user, 'profile') and request.user.profile.role == 'student':
        messages.error(request, "Access denied. Students are not allowed to view the hostel student directory.")
        return redirect('students:home')
        
    from django.db.models import Q
    
    students = Student.objects.filter(student_type='hostel', is_active=True).exclude(alumni_record__isnull=False)
    
    search_query = request.GET.get('search', '')
    if search_query:
        students = students.filter(
            Q(student_id__icontains=search_query) |
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query)
        )
        
    # Calculate current status for each
    away_student_ids = set(HostelMovement.objects.filter(is_returned=False).values_list('student_id', flat=True))
    
    for st in students:
        st.is_away = st.id in away_student_ids
        
    context = {
        'students': students,
        'search_query': search_query
    }
    return render(request, 'students/hostel_student_list.html', context)


@role_required(['admin', 'ntstaff', 'student'])
def hostel_student_detail_view(request, pk):
    """Detailed profile for a specific hostel student"""
    # Data isolation for students
    if hasattr(request.user, 'profile') and request.user.profile.role == 'student':
        if not request.user.profile.student_record or request.user.profile.student_record.id != int(pk):
            from django.contrib import messages
            messages.error(request, "You do not have permission to view other students' hostel details.")
            return redirect('students:home')
    from datetime import datetime, date
    
    student = get_object_or_404(Student, pk=pk, student_type='hostel')
    movements = HostelMovement.objects.filter(student=student).order_by('-departure_date', '-departure_time')
    
    # Calculate historical away days
    total_away_days = 0
    today = timezone.now().date()
    
    for movement in movements:
        end_date = movement.arrival_date if movement.is_returned else today
        start_date = movement.departure_date
        total_away_days += (end_date - start_date).days
        
    # Enrollment baseline (when they joined)
    enroll_date = student.created_at.date()
    total_days_since_enroll = (today - enroll_date).days
    total_present_days = max(0, total_days_since_enroll - total_away_days)
    
    # Check specific date logic
    check_date_str = request.GET.get('check_date')
    check_status = None
    
    if check_date_str:
        try:
            check_date = datetime.strptime(check_date_str, '%Y-%m-%d').date()
            # If the date is before enrollment or after today
            if check_date < enroll_date:
                check_status = "Not Enrolled Yet"
            elif check_date > today:
                check_status = "Future Date"
            else:
                # Find if any movement perfectly covers this date
                was_away = False
                for movement in movements:
                    m_end = movement.arrival_date if movement.is_returned else today
                    if movement.departure_date <= check_date <= m_end:
                        was_away = True
                        break
                check_status = "Away" if was_away else "Present"
        except ValueError:
            pass
            
    is_currently_away = movements.filter(is_returned=False).exists()
    
    context = {
        'student': student,
        'movements': movements,
        'total_away_days': total_away_days,
        'total_present_days': total_present_days,
        'is_currently_away': is_currently_away,
        'check_date_str': check_date_str,
        'check_status': check_status
    }
    return render(request, 'students/hostel_student_detail.html', context)


@role_required(['admin'])
def section_create(request):
    """Create a new section"""
    if request.method == 'POST':
        form = SectionForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Section created successfully.')
            return redirect('students:section_list')
    else:
        form = SectionForm()
    
    context = {
        'form': form,
        'title': 'Create Section',
        'is_update': False
    }
    return render(request, 'students/section_form.html', context)


@role_required(['admin'])
def section_update(request, pk):
    """Update an existing section"""
    section = get_object_or_404(Section, pk=pk)
    if request.method == 'POST':
        form = SectionForm(request.POST, instance=section)
        if form.is_valid():
            form.save()
            messages.success(request, 'Section updated successfully.')
            return redirect('students:section_list')
    else:
        form = SectionForm(instance=section)
    
    context = {
        'form': form,
        'title': 'Update Section',
        'is_update': True
    }
    return render(request, 'students/section_form.html', context)


@role_required(['admin'])
def section_delete(request, pk):
    """Delete an existing section"""
    section = get_object_or_404(Section, pk=pk)
    section.delete()
    messages.success(request, 'Section deleted successfully.')
    return redirect('students:section_list')

# ==========================================
# ACADEMIC YEAR MANAGEMENT
# ==========================================
from .forms import AcademicYearForm, SubjectForm

@role_required(['admin'])
def academic_year_list(request):
    """List all academic years"""
    academic_years = AcademicYear.objects.all().order_by('-start_date')
    context = {'academic_years': academic_years}
    return render(request, 'students/academic_year_list.html', context)


@role_required(['admin'])
def academic_year_create(request):
    """Create a new academic year"""
    if request.method == 'POST':
        form = AcademicYearForm(request.POST)
        if form.is_valid():
            year = form.save()
            # If this is set as active, deactivate others
            if year.is_active:
                AcademicYear.objects.exclude(id=year.id).update(is_active=False)
            messages.success(request, 'Academic Year created successfully.')
            return redirect('students:academic_year_list')
    else:
        form = AcademicYearForm()
    
    context = {
        'form': form,
        'title': 'Create Academic Year',
        'is_update': False
    }
    return render(request, 'students/academic_year_form.html', context)


@role_required(['admin'])
def academic_year_update(request, pk):
    """Update an existing academic year"""
    year = get_object_or_404(AcademicYear, pk=pk)
    if request.method == 'POST':
        form = AcademicYearForm(request.POST, instance=year)
        if form.is_valid():
            updated_year = form.save()
            # If this is set as active, deactivate others
            if updated_year.is_active:
                AcademicYear.objects.exclude(id=updated_year.id).update(is_active=False)
            messages.success(request, 'Academic Year updated successfully.')
            return redirect('students:academic_year_list')
    else:
        form = AcademicYearForm(instance=year)
    
    context = {
        'form': form,
        'title': 'Update Academic Year',
        'is_update': True
    }
    return render(request, 'students/academic_year_form.html', context)


# ==========================================
# SUBJECT MANAGEMENT
# ==========================================

@role_required(['admin'])
def subject_list(request):
    """List all subjects with filtering"""
    subjects = Subject.objects.all().order_by('grade', 'subject_type', 'name')
    
    # Filters
    grade = request.GET.get('grade')
    subject_type = request.GET.get('subject_type')
    
    if grade:
        subjects = subjects.filter(grade=grade)
    if subject_type:
        subjects = subjects.filter(subject_type=subject_type)
        
    # Get all unique grades for the filter dropdown
    grades = list(Subject.objects.values_list('grade', flat=True).distinct().order_by('grade'))
    
    context = {
        'subjects': subjects,
        'grades': grades,
        'current_filters': {
            'grade': grade,
            'subject_type': subject_type
        }
    }
    return render(request, 'students/subject_list.html', context)


@role_required(['admin'])
def subject_create(request):
    """Create a new subject"""
    if request.method == 'POST':
        form = SubjectForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Subject created successfully.')
            return redirect('students:subject_list')
    else:
        form = SubjectForm()
        
    context = {
        'form': form,
        'title': 'Create Subject',
        'is_update': False
    }
    return render(request, 'students/subject_form.html', context)


@role_required(['admin'])
def subject_update(request, pk):
    """Update an existing subject"""
    subject = get_object_or_404(Subject, pk=pk)
    if request.method == 'POST':
        form = SubjectForm(request.POST, instance=subject)
        if form.is_valid():
            form.save()
            messages.success(request, 'Subject updated successfully.')
            return redirect('students:subject_list')
    else:
        form = SubjectForm(instance=subject)
        
    context = {
        'form': form,
        'title': 'Update Subject',
        'is_update': True
    }
    return render(request, 'students/subject_form.html', context)


# --- Enquiry Views ---

def enquiry_create_view(request):
    """Public or admin view to submit a new enquiry"""
    from .forms import EnquiryForm
    from .models import Division, AcademicYear
    import json
    
    if request.method == 'POST':
        form = EnquiryForm(request.POST)
        if form.is_valid():
            enquiry = form.save(commit=False)
            
            # Auto-assign the active academic year for the selected section (or globally)
            active_year = AcademicYear.objects.filter(is_active=True).first()
            if active_year:
                enquiry.academic_year = active_year
                
            enquiry.save()
            return redirect('students:enquiry_success', application_number=enquiry.application_number)
    else:
        form = EnquiryForm()

    # Pass division mapping for the dependent course dropdown
    divisions = list(Division.objects.values('id', 'name', 'section_id'))
    divisions_json = json.dumps(divisions)
        
    return render(request, 'students/enquiry_form.html', {
        'form': form,
        'divisions_json': divisions_json
    })

def enquiry_success_view(request, application_number):
    """Public page shown after successfully submitting an enquiry"""
    from .models import Enquiry
    enquiry = get_object_or_404(Enquiry, application_number=application_number)
    return render(request, 'students/enquiry_success.html', {
        'application_number': application_number,
        'enquiry': enquiry
    })

def enquiry_status_view(request):
    """Public page to check enquiry status using application number"""
    from .models import Enquiry
    enquiry = None
    searched = False
    
    # Check for both form submission (POST) or direct link (GET parameter)
    app_number = None
    if request.method == 'POST':
        app_number = request.POST.get('application_number', '').strip()
    else:
        app_number = request.GET.get('app_no', '').strip()
        
    if app_number:
        searched = True
        enquiry = Enquiry.objects.filter(application_number__iexact=app_number).first()
        if not enquiry:
            messages.error(request, 'Application not found. Please check your application number.')
            
    from .models import GlobalSettings
    global_settings = GlobalSettings.load()
            
    return render(request, 'students/enquiry_status.html', {
        'enquiry': enquiry,
        'searched': searched,
        'global_settings': global_settings
    })


def enquiry_list_view(request):
    """Dashboard to list all enquiries in tabbed tables"""
    from .models import Enquiry, GlobalSettings
    import urllib.parse
    
    # Base query ordered by newest
    all_enquiries = Enquiry.objects.all().order_by('-created_at')
    global_settings = GlobalSettings.load()
    template = global_settings.whatsapp_message_template
    
    # Pre-compute WhatsApp URLs per enquiry to keep templates clean
    for enquiry in all_enquiries:
        status_url = request.build_absolute_uri(f"/students/enquiries/status/?app_no={enquiry.application_number}")
        message = template.format(
            name=enquiry.name,
            app_no=enquiry.application_number,
            status_link=status_url
        )
        enquiry.whatsapp_url = f"https://wa.me/{urllib.parse.urlencode({'phone': enquiry.phone})[6:]}?text={urllib.parse.quote(message)}"

    # Separate querysets for tabs
    pending_enquiries = [e for e in all_enquiries if e.status == 'Pending']
    received_enquiries = [e for e in all_enquiries if e.status == 'Received']
    token_enquiries = [e for e in all_enquiries if e.status == 'Token Generated']
    enrolled_enquiries = [e for e in all_enquiries if e.status == 'Enrolled']

    return render(request, 'students/enquiry_list.html', {
        'pending_enquiries': pending_enquiries,
        'received_enquiries': received_enquiries,
        'token_enquiries': token_enquiries,
        'enrolled_enquiries': enrolled_enquiries,
        'global_settings': global_settings
    })


@role_required(['admin', 'teacher', 'ntstaff'])
def enquiry_print_list_view(request):
    """Renders a clean, print-friendly list of all enquiries"""
    from .models import Enquiry
    enquiries = Enquiry.objects.all().order_by('-created_at')
    return render(request, 'students/enquiry_print_list.html', {
        'enquiries': enquiries,
        'today': timezone.now()
    })

def enquiry_generate_token_view(request, pk):
    """Generates the next sequential token for the day for the specific enquiry"""
    from .models import Enquiry
    from django.db.models import Max
    
    enquiry = get_object_or_404(Enquiry, pk=pk)
    
    if request.method == 'POST':
        if enquiry.status == 'Pending' or enquiry.status == 'Received':
            # Generate Token Logic
            today = timezone.now().date()
            # Find max token for today
            max_token = Enquiry.objects.filter(
                updated_at__date=today, 
                status='Token Generated'
            ).aggregate(Max('token_number'))['token_number__max']
            
            next_token = (max_token or 0) + 1
            
            enquiry.token_number = next_token
            enquiry.status = 'Token Generated'
            enquiry.save()
            messages.success(request, f'Token #{next_token} generated for {enquiry.name}.')
    
    return redirect('students:enquiry_list')

@role_required(['admin', 'teacher', 'ntstaff'])
def enquiry_mark_received(request, pk):
    """Marks a pending enquiry as Received"""
    from .models import Enquiry
    enquiry = get_object_or_404(Enquiry, pk=pk)
    
    if request.method == 'POST' and enquiry.status == 'Pending':
        enquiry.status = 'Received'
        enquiry.save()
        messages.success(request, f'Application {enquiry.application_number} marked as Received.')
        
    return redirect('students:enquiry_list')

@role_required(['admin', 'teacher', 'ntstaff'])
def enquiry_set_global_interview_date(request):
    """Sets the global, common interview date for all enquiries"""
    from .models import GlobalSettings
    
    if request.method == 'POST':
        form_type = request.POST.get('form_type')
        
        try:
            settings = GlobalSettings.load()
            
            if form_type == 'date_form':
                date_str = request.POST.get('global_interview_date')
                settings.common_interview_date = date_str if date_str else None
                settings.save()
                messages.success(request, f'Global interview date updated successfully.')
                
            elif form_type == 'whatsapp_form':
                msg_template = request.POST.get('whatsapp_message_template')
                if msg_template:
                    settings.whatsapp_message_template = msg_template
                    settings.save()
                    messages.success(request, f'WhatsApp message template updated successfully.')
                    
        except Exception as e:
            messages.error(request, f'Failed to update settings. Error: {e}')
                
    return redirect('students:enquiry_list')

def enquiry_token_print_view(request, pk):
    """View to print a generated token"""
    from .models import Enquiry
    enquiry = get_object_or_404(Enquiry, pk=pk)
    return render(request, 'students/enquiry_token_print.html', {'enquiry': enquiry})


def enquiry_enroll_view(request, pk):
    """One-click transition from Enquiry to Student profile"""
    from .models import Enquiry, Student
    
    enquiry = get_object_or_404(Enquiry, pk=pk)
    
    if enquiry.status == 'Enrolled':
        messages.info(request, 'This enquiry is already enrolled.')
        return redirect('students:enquiry_list')

    if request.method == 'POST':
        # Create a new student profile and pre-fill fields
        # Using a temporary unique ID generation for the newly enrolled student
        import random
        temp_id = f"TEMP-{random.randint(1000, 9999)}"
        
        # We assume the name has first and last components, or store everything in first_name
        name_parts = enquiry.name.split(' ', 1)
        first_name = name_parts[0]
        last_name = name_parts[1] if len(name_parts) > 1 else ''
        
        # Create student profile
        new_student = Student.objects.create(
            student_id=temp_id,
            first_name=first_name,
            last_name=last_name,
            phone=enquiry.phone,
            address=enquiry.district
        )
        
        # Update Enquiry status
        enquiry.status = 'Enrolled'
        enquiry.save()
        
        # Check if academic year is available from enquiry, otherwise fallback
        from .models import AcademicYear, Enrollment
        year = enquiry.academic_year
        if not year:
            year = AcademicYear.objects.filter(is_active=True).first()

        if year:
            Enrollment.objects.create(
                student=new_student,
                academic_year=year,
                section=enquiry.section,
                division=enquiry.course,
                grade=None  # Admin must set proper Grade on edit page
            )

        messages.success(request, f'Enquiry {enquiry.name} enrolled successfully! Please assign their proper Student ID and Academic Enrollment below.')
        # Redirect to the student edit form to complete other fields
        return redirect('students:student_edit', pk=new_student.pk)
        
    return redirect('students:enquiry_list')

@role_required(['admin', 'teacher', 'ntstaff'])
def enquiry_update_view(request, pk):
    """View to edit an existing enquiry"""
    from .models import Enquiry, Division
    from .forms import EnquiryForm
    import json
    
    enquiry = get_object_or_404(Enquiry, pk=pk)
    
    if request.method == 'POST':
        form = EnquiryForm(request.POST, instance=enquiry)
        if form.is_valid():
            form.save()
            messages.success(request, f'Enquiry for {enquiry.name} updated successfully.')
            return redirect('students:enquiry_list')
    else:
        form = EnquiryForm(instance=enquiry)

    # Pass division mapping for the dependent course dropdown
    divisions = list(Division.objects.values('id', 'name', 'section_id'))
    divisions_json = json.dumps(divisions)
        
    return render(request, 'students/enquiry_edit.html', {
        'form': form,
        'enquiry': enquiry,
        'divisions_json': divisions_json
    })

@role_required(['admin', 'teacher', 'ntstaff'])
def enquiry_delete_view(request, pk):
    """View to delete an enquiry"""
    from .models import Enquiry
    enquiry = get_object_or_404(Enquiry, pk=pk)
    
    if request.method == 'POST':
        name = enquiry.name
        enquiry.delete()
        messages.success(request, f'Enquiry for {name} deleted successfully.')
        return redirect('students:enquiry_list')
        
    return render(request, 'students/enquiry_confirm_delete.html', {'enquiry': enquiry})


# --- GRADE CRUD ---

@role_required(['admin'])
def grade_list(request):
    grades = Grade.objects.all()
    return render(request, 'students/grade_list.html', {'grades': grades})

@role_required(['admin'])
def grade_create(request):
    if request.method == 'POST':
        form = GradeForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Grade created successfully.')
            return redirect('students:grade_list')
    else:
        form = GradeForm()
    return render(request, 'students/grade_form.html', {'form': form})

@role_required(['admin'])
def grade_update(request, pk):
    grade = get_object_or_404(Grade, pk=pk)
    if request.method == 'POST':
        form = GradeForm(request.POST, instance=grade)
        if form.is_valid():
            form.save()
            messages.success(request, 'Grade updated successfully.')
            return redirect('students:grade_list')
    else:
        form = GradeForm(instance=grade)
    return render(request, 'students/grade_form.html', {'form': form, 'is_update': True})

@role_required(['admin'])
def grade_delete(request, pk):
    grade = get_object_or_404(Grade, pk=pk)
    if request.method == 'POST':
        grade.delete()
        messages.success(request, 'Grade deleted successfully.')
        return redirect('students:grade_list')
    return render(request, 'students/grade_confirm_delete.html', {'grade': grade})

# --- DIVISION CRUD ---

@role_required(['admin'])
def division_list(request):
    divisions = Division.objects.all()
    return render(request, 'students/division_list.html', {'divisions': divisions})

@role_required(['admin'])
def division_create(request):
    if request.method == 'POST':
        form = DivisionForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Division created successfully.')
            return redirect('students:division_list')
    else:
        form = DivisionForm()
    return render(request, 'students/division_form.html', {'form': form})

@role_required(['admin'])
def division_update(request, pk):
    division = get_object_or_404(Division, pk=pk)
    if request.method == 'POST':
        form = DivisionForm(request.POST, instance=division)
        if form.is_valid():
            form.save()
            messages.success(request, 'Division updated successfully.')
            return redirect('students:division_list')
    else:
        form = DivisionForm(instance=division)
    return render(request, 'students/division_form.html', {'form': form, 'is_update': True})

@role_required(['admin'])
def division_delete(request, pk):
    division = get_object_or_404(Division, pk=pk)
    if request.method == 'POST':
        division.delete()
        messages.success(request, 'Division deleted successfully.')
        return redirect('students:division_list')
    return render(request, 'students/division_confirm_delete.html', {'division': division})

# --- SUBJECT CRUD ---

@role_required(['admin'])
def subject_list(request):
    subjects = Subject.objects.select_related('grade', 'division', 'section').all()
    return render(request, 'students/subject_list.html', {'subjects': subjects})

@role_required(['admin'])
def subject_create(request):
    if request.method == 'POST':
        form = SubjectForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Subject created successfully.')
            return redirect('students:subject_list')
    else:
        form = SubjectForm()
    return render(request, 'students/subject_form.html', {'form': form})

@role_required(['admin'])
def subject_update(request, pk):
    subject = get_object_or_404(Subject, pk=pk)
    if request.method == 'POST':
        form = SubjectForm(request.POST, instance=subject)
        if form.is_valid():
            form.save()
            messages.success(request, 'Subject updated successfully.')
            return redirect('students:subject_list')
    else:
        form = SubjectForm(instance=subject)
    return render(request, 'students/subject_form.html', {'form': form, 'is_update': True})

@role_required(['admin'])
def subject_delete(request, pk):
    subject = get_object_or_404(Subject, pk=pk)
    if request.method == 'POST':
        subject.delete()
        messages.success(request, 'Subject deleted successfully.')
        return redirect('students:subject_list')
    return render(request, 'students/subject_confirm_delete.html', {'subject': subject})

from django.contrib.auth.models import User
from .forms import UserManageForm
from .models import UserProfile

@role_required(['admin'])
def user_list(request):
    users = User.objects.select_related('profile').all().order_by('-date_joined')
    return render(request, 'students/user_list.html', {'users': users})

@role_required(['admin'])
def user_create(request):
    if request.method == 'POST':
        form = UserManageForm(request.POST)
        if form.is_valid():
            user = form.save(commit=False)
            password = form.cleaned_data.get('password')
            if password:
                user.set_password(password)
            user.save()
            
            # The signal might create a default profile, so let's update it or create it
            profile, created = UserProfile.objects.get_or_create(user=user)
            profile.role = form.cleaned_data.get('role')
            profile.student_record = form.cleaned_data.get('student_record')
            profile.save()
            
            messages.success(request, f'User {user.username} created successfully.')
            return redirect('students:user_list')
    else:
        form = UserManageForm()
    
    return render(request, 'students/user_form.html', {'form': form, 'title': 'Create User'})

@role_required(['admin'])
def user_update(request, pk):
    user = get_object_or_404(User, pk=pk)
    profile = getattr(user, 'profile', None)
    
    if request.method == 'POST':
        form = UserManageForm(request.POST, instance=user)
        if form.is_valid():
            updated_user = form.save(commit=False)
            password = form.cleaned_data.get('password')
            if password:
                updated_user.set_password(password)
            updated_user.save()
            
            if not profile:
                profile = UserProfile.objects.create(user=updated_user)
                
            profile.role = form.cleaned_data.get('role')
            profile.student_record = form.cleaned_data.get('student_record')
            profile.save()
            
            messages.success(request, f'User {updated_user.username} updated successfully.')
            return redirect('students:user_list')
    else:
        initial = {}
        if profile:
            initial['role'] = profile.role
            initial['student_record'] = profile.student_record
        form = UserManageForm(instance=user, initial=initial)
        
    return render(request, 'students/user_form.html', {'form': form, 'title': 'Edit User'})

@role_required(['admin'])
def user_delete(request, pk):
    user_obj = get_object_or_404(User, pk=pk)
    if user_obj.is_superuser and request.user.pk != user_obj.pk:
        messages.error(request, "You cannot delete other superusers.")
        return redirect('students:user_list')
        
    if request.method == 'POST':
        user_obj.delete()
        messages.success(request, 'User deleted successfully.')
        return redirect('students:user_list')
    return render(request, 'students/user_confirm_delete.html', {'user_obj': user_obj})

from .models import Alumni
from .forms import AlumniTransferForm, AlumniForm

@login_required
@role_required(['admin', 'teacher'])
def alumni_list(request):
    """List all registered alumni"""
    alumni_records = Alumni.objects.select_related('student').all()
    
    # Generic search
    search_query = request.GET.get('search', '')
    if search_query:
        alumni_records = alumni_records.filter(
            Q(student__first_name__icontains=search_query) |
            Q(student__last_name__icontains=search_query) |
            Q(student__student_id__icontains=search_query) |
            Q(graduation_year__icontains=search_query)
        )
        
    context = {
        'alumni_records': alumni_records,
        'search_query': search_query,
        'page_title': 'Alumni Registry'
    }
    return render(request, 'students/alumni_list.html', context)

@login_required
@role_required(['admin'])
def transfer_to_alumni(request, student_id):
    """Transfer an active student to alumni status"""
    student = get_object_or_404(Student, id=student_id)
    
    # Check if already alumni
    if hasattr(student, 'alumni_record'):
        messages.warning(request, f"{student.full_name} is already registered as an Alumni.")
        return redirect('students:alumni_list')
        
    if request.method == 'POST':
        form = AlumniTransferForm(request.POST)
        if form.is_valid():
            alumni = form.save(commit=False)
            alumni.student = student
            alumni.save()
            
            # Deactivate student immediately upon transfer
            student.is_active = False
            student.save()
            
            messages.success(request, f"Successfully transferred {student.full_name} to the Alumni Registry.")
            return redirect('students:alumni_list')
    else:
        # Default graduation year to current year
        import datetime
        form = AlumniTransferForm(initial={'graduation_year': str(datetime.date.today().year)})
        
    context = {
        'form': form,
        'student': student,
        'page_title': f'Transfer {student.full_name} to Alumni'
    }
    return render(request, 'students/alumni_form.html', context)

@login_required
@role_required(['admin'])
def alumni_update(request, pk):
    alumni = get_object_or_404(Alumni, pk=pk)
    
    if request.method == 'POST':
        form = AlumniForm(request.POST, instance=alumni)
        if form.is_valid():
            form.save()
            messages.success(request, f"Alumni record for {alumni.student.full_name} updated successfully.")
            return redirect('students:alumni_list')
    else:
        form = AlumniForm(instance=alumni)
        
    context = {
        'form': form,
        'alumni': alumni,
        'is_update': True,
        'page_title': f'Update Alumni: {alumni.student.full_name}'
    }
    return render(request, 'students/alumni_form.html', context)

@login_required
@role_required(['admin'])
def alumni_delete(request, pk):
    alumni = get_object_or_404(Alumni, pk=pk)
    if request.method == 'POST':
        student = alumni.student
        reactivate = request.POST.get('reactivate') == 'on'
        
        alumni.delete()
        
        if reactivate:
            student.is_active = True
            student.save()
            messages.success(request, f"Alumni record deleted and {student.full_name} has been restored to active student status.")
        else:
            messages.success(request, f"Alumni record for {student.full_name} deleted successfully.")
            
        return redirect('students:alumni_list')
        
    return render(request, 'students/alumni_confirm_delete.html', {'alumni': alumni})

from .models import Enrollment, AcademicYear, Grade

@login_required
@role_required(['admin'])
def alumni_bulk_transfer(request):
    """Bulk transfer an entire class to alumni registry"""
    active_year = AcademicYear.objects.filter(is_active=True).first()
    grades = Grade.objects.all().order_by('name')
    
    if request.method == 'POST':
        grade_id = request.POST.get('grade')
        graduation_year = request.POST.get('graduation_year')
        remarks = request.POST.get('remarks', f'Bulk transferred at end of {active_year.name if active_year else "term"}')
        
        if grade_id and graduation_year:
            # Get all active students enrolled in this grade
            enrollments = Enrollment.objects.filter(
                academic_year=active_year, 
                grade_id=grade_id,
                student__is_active=True
            ).select_related('student')
            
            transferred_count = 0
            for enrollment in enrollments:
                student = enrollment.student
                # Avoid duplicates
                if not hasattr(student, 'alumni_record'):
                    Alumni.objects.create(
                        student=student,
                        graduation_year=graduation_year,
                        remarks=remarks,
                        current_status='Graduated (Bulk Transfer)'
                    )
                    student.is_active = False
                    student.save()
                    transferred_count += 1
            
            if transferred_count > 0:
                messages.success(request, f"Successfully created alumni records for {transferred_count} students.")
            else:
                messages.warning(request, "No active students found in that class without an existing alumni record.")
                
            return redirect('students:alumni_list')
        else:
            messages.error(request, "Grade and Graduation Year are required for bulk transfer.")
            
    import datetime
    current_year = str(datetime.date.today().year)
    
    context = {
        'page_title': 'Bulk Transfer Class to Alumni',
        'grades': grades,
        'current_year': current_year,
        'active_year': active_year
    }
    return render(request, 'students/alumni_bulk_transfer.html', context)


@login_required
@role_required(['admin'])
def alumni_bulk_restore(request):
    """Bulk restore/revert an entire class from alumni registry back to active roster"""
    active_year = AcademicYear.objects.filter(is_active=True).first()
    grades = Grade.objects.all().order_by('name')
    
    if request.method == 'POST':
        grade_id = request.POST.get('grade')
        graduation_year = request.POST.get('graduation_year')
        
        if grade_id and graduation_year:
            # Find alumni matching this grade in the active year and the provided graduation year
            alumni_records = Alumni.objects.filter(
                graduation_year=graduation_year,
                student__enrollments__academic_year=active_year,
                student__enrollments__grade_id=grade_id
            ).distinct()
            
            restored_count = 0
            for alumni in alumni_records:
                student = alumni.student
                alumni.delete()
                student.is_active = True
                student.save()
                restored_count += 1
            
            if restored_count > 0:
                messages.success(request, f"Successfully restored {restored_count} students back to active rosters.")
            else:
                messages.warning(request, "No matching alumni records found to restore.")
                
            return redirect('students:alumni_list')
        else:
            messages.error(request, "Grade and Graduation Year are required for bulk restore.")
            
    import datetime
    current_year = str(datetime.date.today().year)
    
    context = {
        'page_title': 'Bulk Restore Class from Alumni',
        'grades': grades,
        'current_year': current_year,
        'active_year': active_year
    }
    return render(request, 'students/alumni_bulk_restore.html', context)


# --- Job Vacancy Views ---
from .models import JobOpening, JobApplication
from .forms import JobOpeningForm, JobApplicationForm

def career_page(request):
    """Public page listing job vacancies"""
    today = date.today()
    job_openings = JobOpening.objects.filter(is_active=True).filter(
        Q(validity__gte=today) | Q(validity__isnull=True)
    )
    return render(request, 'students/career.html', {'job_openings': job_openings})

def job_apply(request, job_id):
    """Public job application form"""
    job = get_object_or_404(JobOpening, id=job_id, is_active=True)
    if request.method == 'POST':
        form = JobApplicationForm(request.POST, request.FILES)
        if form.is_valid():
            application = form.save(commit=False)
            application.job = job
            application.save()
            return redirect('students:job_success', application_number=application.application_number)
    else:
        form = JobApplicationForm()
    
    return render(request, 'students/job_application_form.html', {'form': form, 'job': job})

def job_success(request, application_number):
    """Success page after job application"""
    application = get_object_or_404(JobApplication, application_number=application_number)
    return render(request, 'students/job_success.html', {'application': application})

@role_required(['admin'])
def job_vacancy_list_admin(request):
    """Admin view to manage job vacancies"""
    job_openings = JobOpening.objects.all()
    return render(request, 'students/admin/job_vacancy_list.html', {'job_openings': job_openings})

@role_required(['admin'])
def job_vacancy_create(request):
    """Admin view to create a job vacancy"""
    if request.method == 'POST':
        form = JobOpeningForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Job vacancy announced successfully.')
            return redirect('students:job_vacancy_list_admin')
    else:
        form = JobOpeningForm()
    return render(request, 'students/admin/job_opening_form.html', {'form': form, 'title': 'Announce New Job'})

@role_required(['admin'])
def job_vacancy_update(request, pk):
    """Admin view to update a job vacancy"""
    job = get_object_or_404(JobOpening, pk=pk)
    if request.method == 'POST':
        form = JobOpeningForm(request.POST, instance=job)
        if form.is_valid():
            form.save()
            messages.success(request, 'Job vacancy updated successfully.')
            return redirect('students:job_vacancy_list_admin')
    else:
        form = JobOpeningForm(instance=job)
    return render(request, 'students/admin/job_opening_form.html', {'form': form, 'title': 'Edit Job Vacancy'})

@role_required(['admin'])
def job_vacancy_delete(request, pk):
    """Admin view to delete a job vacancy"""
    job = get_object_or_404(JobOpening, pk=pk)
    if request.method == 'POST':
        job.delete()
        messages.success(request, 'Job vacancy deleted.')
        return redirect('students:job_vacancy_list_admin')
    return render(request, 'students/admin/job_opening_confirm_delete.html', {'job': job})

@role_required(['admin'])
def job_application_list_admin(request):
    """Admin view to list all job applications"""
    applications = JobApplication.objects.all()
    return render(request, 'students/admin/job_application_list.html', {'applications': applications})


def student_results_public_lookup(request):
    """Public portal to search and view exam results by student register number"""
    from django.contrib import messages
    from django.db.models import Sum, Q
    
    # Get all published exams
    published_exams = ExamType.objects.filter(is_published=True).order_by('order', 'name')
    
    student_id = request.GET.get('student_id', '').strip()
    selected_exam_id = request.GET.get('exam_type_id', '').strip()
    
    student = None
    enrollment = None
    results_by_exam = {}
    
    if student_id:
        try:
            student = Student.objects.get(student_id__iexact=student_id, is_active=True)
            
            # Find student enrollment (prefer active academic year)
            active_year = AcademicYear.objects.filter(is_active=True).first()
            if active_year:
                enrollment = Enrollment.objects.filter(student=student, academic_year=active_year).select_related('grade', 'division', 'section').first()
            
            if not enrollment:
                enrollment = Enrollment.objects.filter(student=student).select_related('grade', 'division', 'section').order_by('-academic_year__start_date').first()
                
            # Filter published exams by student's section if exam is section-specific
            exams_to_query = published_exams
            if enrollment and enrollment.section:
                exams_to_query = exams_to_query.filter(Q(section=enrollment.section) | Q(section__isnull=True))
                
            if selected_exam_id:
                try:
                    selected_exam_id_int = int(selected_exam_id)
                    exams_to_query = exams_to_query.filter(id=selected_exam_id_int)
                except ValueError:
                    pass
                    
            for exam in exams_to_query:
                # Fetch mark entries
                mark_entries = MarkEntry.objects.filter(
                    student=student,
                    exam_type=exam
                ).select_related('subject').order_by('subject__subject_type', 'subject__name')
                
                if mark_entries.exists():
                    total_obtained = sum(entry.marks_obtained for entry in mark_entries)
                    total_max = sum(entry.max_marks for entry in mark_entries)
                    overall_percentage = (total_obtained / total_max * 100) if total_max > 0 else 0
                    
                    # Calculate overall grade letter
                    pct = float(overall_percentage)
                    if pct >= 90:
                        overall_grade = 'A+'
                    elif pct >= 80:
                        overall_grade = 'A'
                    elif pct >= 70:
                        overall_grade = 'B+'
                    elif pct >= 60:
                        overall_grade = 'B'
                    elif pct >= 50:
                        overall_grade = 'C+'
                    elif pct >= 40:
                        overall_grade = 'C'
                    else:
                        overall_grade = 'F'
                        
                    # Split into Hadiya and Division subjects
                    hadiya_marks = [e for e in mark_entries if e.subject.subject_type == 'hadiya']
                    division_marks = [e for e in mark_entries if e.subject.subject_type == 'division']
                    
                    results_by_exam[exam] = {
                        'mark_entries': mark_entries,
                        'hadiya_marks': hadiya_marks,
                        'division_marks': division_marks,
                        'total_obtained': total_obtained,
                        'total_max': total_max,
                        'percentage': overall_percentage,
                        'overall_grade': overall_grade,
                    }
                    
            if not results_by_exam and selected_exam_id:
                messages.warning(request, f"No results found for student '{student_id}' in the selected exam.")
            elif not results_by_exam:
                messages.warning(request, f"No published exam results found for student '{student_id}'.")
                
        except Student.DoesNotExist:
            messages.error(request, f"Invalid Register Number: '{student_id}' not found.")
            
    context = {
        'published_exams': published_exams,
        'student_id': student_id,
        'selected_exam_id': selected_exam_id,
        'student': student,
        'enrollment': enrollment,
        'results_by_exam': results_by_exam,
    }
    return render(request, 'students/results_lookup.html', context)


@role_required(['admin', 'teacher', 'accountant'])
def examination_dashboard(request):
    total_exams = ExamType.objects.count()
    published_exams = ExamType.objects.filter(is_published=True).count()
    draft_exams = total_exams - published_exams
    total_subjects = Subject.objects.count()
    total_mark_entries = MarkEntry.objects.count()
    
    context = {
        'total_exams': total_exams,
        'published_exams': published_exams,
        'draft_exams': draft_exams,
        'total_subjects': total_subjects,
        'total_mark_entries': total_mark_entries,
    }
    return render(request, 'students/examination_dashboard.html', context)


@role_required(['admin', 'teacher'])
def monthly_attendance_select(request):
    """Select Month, Year, Class, and Type for Monthly Attendance Grid"""
    active_year = AcademicYear.objects.filter(is_active=True).first()
    if not active_year:
        messages.error(request, 'Please set an active academic year to mark attendance.')
        return redirect('students:home')
        
    today = date.today()
    month = int(request.GET.get('month', today.month))
    year = int(request.GET.get('year', today.year))
    attendance_type = request.GET.get('type', 'daily')
    period_id = request.GET.get('period')
    activity_id = request.GET.get('activity')
    
    grades = Grade.objects.all().order_by('order', 'name')
    divisions = Division.objects.all()
    periods = Period.objects.all()
    activities = Activity.objects.filter(date__year=year, date__month=month) if month else Activity.objects.none()
    
    import calendar
    months_list = [{'value': i, 'name': calendar.month_name[i]} for i in range(1, 13)]
    
    if request.GET.get('grade'):
        grade_id = request.GET.get('grade')
        division_id = request.GET.get('division') or '0'
        params = [
            f"month={month}",
            f"year={year}",
            f"type={attendance_type}"
        ]
        if period_id: params.append(f"period={period_id}")
        if activity_id: params.append(f"activity={activity_id}")
        
        url = reverse('students:monthly_attendance_grid', kwargs={'grade_id': grade_id, 'division_id': division_id})
        return redirect(f"{url}?{'&'.join(params)}")
        
    context = {
        'grades': grades,
        'divisions': divisions,
        'periods': periods,
        'activities': activities,
        'selected_month': month,
        'selected_year': year,
        'selected_type': attendance_type,
        'selected_period': period_id,
        'selected_activity': activity_id,
        'months': months_list,
        'years': range(today.year - 5, today.year + 6),
    }
    return render(request, 'students/monthly_attendance_select.html', context)


@role_required(['admin', 'teacher'])
def monthly_attendance_grid(request, grade_id, division_id):
    """Monthly Attendance Grid for a class"""
    import calendar
    active_year = AcademicYear.objects.filter(is_active=True).first()
    if not active_year:
        messages.error(request, 'Please set an active academic year to mark attendance.')
        return redirect('students:home')

    grade_obj = get_object_or_404(Grade, id=grade_id)
    actual_division_id = None if division_id == 0 else division_id
    
    # Get enrollments for this class
    enrollments = Enrollment.objects.filter(
        academic_year=active_year, 
        student__is_active=True, 
        grade=grade_obj
    ).select_related('student', 'division', 'section')
    
    if actual_division_id:
        enrollments = enrollments.filter(division_id=actual_division_id)
        division = get_object_or_404(Division, id=actual_division_id)
        class_name = f"{grade_obj.name} - {division.name}"
    else:
        enrollments = enrollments.filter(division__isnull=True)
        class_name = f"{grade_obj.name} - No Division"

    section_name = grade_obj.section.name if grade_obj.section else "No Section"

    # Month and Year parameters
    today = date.today()
    month = int(request.GET.get('month', today.month))
    year = int(request.GET.get('year', today.year))
    attendance_type = request.GET.get('type', 'daily')
    period_id = request.GET.get('period')
    activity_id = request.GET.get('activity')

    # Get all days in the month
    num_days = calendar.monthrange(year, month)[1]
    
    # Get holidays for this month
    from students.models import Holiday
    holidays = Holiday.objects.filter(date__year=year, date__month=month)
    if grade_obj:
        holidays = holidays.filter(models.Q(grades__isnull=True) | models.Q(grades=grade_obj)).distinct()
    holiday_map = {h.date: h.title for h in holidays}
    
    periods = Period.objects.all()
    activities = Activity.objects.filter(date__year=year, date__month=month)
    
    selected_period = None
    if period_id:
        selected_period = get_object_or_404(Period, id=period_id)
        
    selected_activity = None
    if activity_id:
        selected_activity = get_object_or_404(Activity, id=activity_id)

    # Base filter for attendance query
    attendance_filter = {
        'enrollment__in': enrollments,
        'date__year': year,
        'date__month': month,
        'attendance_type': attendance_type,
    }
    if attendance_type == 'period' and period_id:
        attendance_filter['period_id'] = period_id
    elif attendance_type == 'activity' and activity_id:
        attendance_filter['activity_id'] = activity_id
    else:
        attendance_filter['period__isnull'] = True
        attendance_filter['activity__isnull'] = True

    # Get existing attendances in this month for these enrollments
    attendances = Attendance.objects.filter(**attendance_filter)
    
    # Organize attendance by student_id and date
    attendance_map = {}
    for att in attendances:
        if att.student_id not in attendance_map:
            attendance_map[att.student_id] = {}
        attendance_map[att.student_id][att.date] = att.status

    # Identify which days already have some attendance marked
    days_with_attendance = set(attendances.values_list('date', flat=True))

    # Generate dates list with metadata
    dates_meta = []
    for d in range(1, num_days + 1):
        curr_date = date(year, month, d)
        weekday = curr_date.weekday()
        
        is_sunday = weekday == 6
        is_second_saturday = (
            weekday == 5 and
            ((d + date(year, month, 1).weekday()) // 7 + 1) == 2
        )
        is_manual_holiday = curr_date in holiday_map
        is_holiday = is_sunday or is_second_saturday or is_manual_holiday
        holiday_type = (
            holiday_map[curr_date] if is_manual_holiday
            else 'Sunday' if is_sunday
            else 'Second Saturday' if is_second_saturday
            else ''
        )
        
        is_future = curr_date > today
        
        # Pre-select "Mark Day" if it already has attendance, OR if it's today/past and NOT a holiday
        should_mark = curr_date in days_with_attendance or (not is_future and not is_holiday)
        
        dates_meta.append({
            'date': curr_date,
            'day_num': d,
            'day_name': curr_date.strftime('%a'),
            'is_holiday': is_holiday,
            'holiday_type': holiday_type,
            'is_future': is_future,
            'should_mark': should_mark,
        })

    if request.method == 'POST':
        marked_by = request.user.get_full_name() or request.user.username or "Admin"
        updated_days = 0
        
        for dm in dates_meta:
            curr_date = dm['date']
            if dm['is_future']:
                continue
            
            # Check if this day is selected to be marked
            mark_day_key = f"mark_day_{curr_date.isoformat()}"
            should_save = mark_day_key in request.POST
            
            if should_save:
                # Update or create attendance for all students for this day
                for enrollment in enrollments:
                    student = enrollment.student
                    
                    # The checkbox name for the student on this day
                    student_att_key = f"att_{student.id}_{curr_date.isoformat()}"
                    
                    status = 'present' if student_att_key in request.POST else 'absent'
                    
                    filter_kwargs = {
                        'student': student,
                        'enrollment': enrollment,
                        'date': curr_date,
                        'attendance_type': attendance_type,
                    }
                    if attendance_type == 'period':
                        filter_kwargs['period'] = selected_period
                    elif attendance_type == 'activity':
                        filter_kwargs['activity'] = selected_activity
                    else:
                        filter_kwargs['period__isnull'] = True
                        filter_kwargs['activity__isnull'] = True

                    Attendance.objects.update_or_create(
                        defaults={
                            'status': status,
                            'marked_by': marked_by,
                        },
                        **filter_kwargs
                    )
                updated_days += 1
            else:
                # If they UNCHECKED a day that previously had attendance, we delete it
                if curr_date in days_with_attendance:
                    delete_filter = {
                        'enrollment__in': enrollments,
                        'date': curr_date,
                        'attendance_type': attendance_type,
                    }
                    if attendance_type == 'period' and period_id:
                        delete_filter['period_id'] = period_id
                    elif attendance_type == 'activity' and activity_id:
                        delete_filter['activity_id'] = activity_id
                    else:
                        delete_filter['period__isnull'] = True
                        delete_filter['activity__isnull'] = True
                        
                    Attendance.objects.filter(**delete_filter).delete()
                    updated_days += 1

        messages.success(request, f'Attendance saved successfully for {updated_days} days.')
        
        # Redirect back to grid
        params = [
            f"month={month}",
            f"year={year}",
            f"type={attendance_type}"
        ]
        if period_id: params.append(f"period={period_id}")
        if activity_id: params.append(f"activity={activity_id}")
        
        url = reverse('students:monthly_attendance_grid', kwargs={'grade_id': grade_id, 'division_id': division_id})
        return redirect(f"{url}?{'&'.join(params)}")

    months_list = [{'value': i, 'name': calendar.month_name[i]} for i in range(1, 13)]
    grades = Grade.objects.all().order_by('order', 'name')
    divisions = Division.objects.all()

    context = {
        'grade_id': int(grade_id),
        'division_id': int(division_id),
        'class_name': class_name,
        'section_name': section_name,
        'enrollments': enrollments,
        'dates_meta': dates_meta,
        'attendance_map': attendance_map,
        'current_month': month,
        'current_year': year,
        'attendance_type': attendance_type,
        'selected_period': period_id,
        'selected_activity': activity_id,
        'periods': periods,
        'activities': activities,
        'months': months_list,
        'years': range(today.year - 5, today.year + 6),
        'grades': grades,
        'divisions': divisions,
        'today': today,
    }
    return render(request, 'students/monthly_attendance_grid.html', context)


# ==========================================
# Face Recognition Attendance Views
# ==========================================

import json
import math
from django.views.decorators.http import require_POST
from django.contrib import messages
from django.shortcuts import get_object_or_404, render, redirect
from django.http import JsonResponse
from django.utils import timezone
from .models import Student, StudentFace, Attendance, Enrollment, Grade, Division, Section, AcademicYear

@role_required(['admin', 'teacher', 'ntstaff'])
def register_face(request, pk):
    """Render face registration page for a specific student"""
    student = get_object_or_404(Student, id=pk)
    face_profile = getattr(student, 'face_profile', None)
    context = {
        'student': student,
        'face_profile': face_profile,
    }
    return render(request, 'students/register_face.html', context)

@role_required(['admin', 'teacher', 'ntstaff'])
@require_POST
def save_face_profile(request, pk):
    """Save face embedding vector and base64 snapshot for a student"""
    student = get_object_or_404(Student, id=pk)
    try:
        data = json.loads(request.body)
        embedding = data.get('embedding')
        photo = data.get('photo') # base64 image data string
        
        if not embedding or not isinstance(embedding, list) or len(embedding) != 128:
            return JsonResponse({'success': False, 'error': 'Invalid embedding data. Must be a 128-float array.'}, status=400)
            
        if not photo:
            return JsonResponse({'success': False, 'error': 'Photo snapshot is required.'}, status=400)
            
        # Update or create face profile
        face_profile, created = StudentFace.objects.update_or_create(
            student=student,
            defaults={
                'embedding': json.dumps(embedding),
                'photo': photo
            }
        )
        return JsonResponse({'success': True, 'message': 'Face profile saved successfully!'})
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Malformed JSON body.'}, status=400)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@role_required(['admin', 'teacher', 'ntstaff'])
def face_attendance_scanner(request):
    """Renders the webcam scanner dashboard for marking attendance by face"""
    grades = Grade.objects.all().order_by('order', 'name')
    divisions = Division.objects.all()
    sections = Section.objects.all().order_by('order', 'name')
    
    context = {
        'grades': grades,
        'divisions': divisions,
        'sections': sections,
    }
    return render(request, 'students/face_attendance_scanner.html', context)

@role_required(['admin', 'teacher', 'ntstaff'])
@require_POST
def mark_face_attendance_ajax(request):
    """Compare scanned face embedding against database and mark attendance"""
    try:
        data = json.loads(request.body)
        input_emb = data.get('embedding')
        
        grade_id = data.get('grade_id')
        division_id = data.get('division_id')
        section_id = data.get('section_id')
        
        if not input_emb or not isinstance(input_emb, list) or len(input_emb) != 128:
            return JsonResponse({'success': False, 'error': 'Invalid face embedding.'}, status=400)
            
        # Filter student faces based on grade/division/section filters to narrow down search if provided
        face_profiles = StudentFace.objects.all().select_related('student')
        
        # Build candidate filter list
        if grade_id or division_id or section_id:
            active_year = AcademicYear.objects.filter(is_active=True).first()
            if active_year:
                enrollments = Enrollment.objects.filter(academic_year=active_year)
                if grade_id:
                    enrollments = enrollments.filter(grade_id=grade_id)
                if division_id:
                    enrollments = enrollments.filter(division_id=division_id)
                if section_id:
                    enrollments = enrollments.filter(section_id=section_id)
                
                candidate_student_ids = enrollments.values_list('student_id', flat=True)
                face_profiles = face_profiles.filter(student_id__in=candidate_student_ids)
        
        best_match = None
        min_distance = 0.5 # Threshold for face-api.js euclidean distance (0.6 is default, 0.5 is safe for security)
        
        for fp in face_profiles:
            try:
                db_emb = json.loads(fp.embedding)
                # Compute Euclidean Distance
                dist = math.sqrt(sum((a - b) ** 2 for a, b in zip(input_emb, db_emb)))
                if dist < min_distance:
                    min_distance = dist
                    best_match = fp
            except Exception:
                continue
                
        if not best_match:
            return JsonResponse({'success': False, 'error': 'Unknown face. Student not found.'}, status=200)
            
        student = best_match.student
        
        # Verify enrollment for active year
        active_year = AcademicYear.objects.filter(is_active=True).first()
        if not active_year:
            return JsonResponse({'success': False, 'error': 'No active academic year is configured.'}, status=400)
            
        enrollment = student.enrollments.filter(academic_year=active_year).first()
        if not enrollment:
            return JsonResponse({'success': False, 'error': f'Student is not enrolled in the active academic year {active_year.name}.'}, status=400)
            
        today = timezone.now().date()
        
        # Check if already marked present
        attendance_exists = Attendance.objects.filter(
            student=student,
            date=today,
            attendance_type='daily'
        ).exists()
        
        if attendance_exists:
            return JsonResponse({
                'success': True,
                'already_marked': True,
                'student_id': student.student_id,
                'name': student.full_name,
                'class': enrollment.class_name,
                'photo': best_match.photo,
                'message': f'{student.full_name} is already marked present today.'
            })
            
        # Mark attendance
        marked_by_user = request.user.get_full_name() or request.user.username
        Attendance.objects.create(
            student=student,
            enrollment=enrollment,
            date=today,
            attendance_type='daily',
            status='present',
            marked_by=marked_by_user,
            remarks='Marked by Face Recognition Scanner'
        )
        
        return JsonResponse({
            'success': True,
            'already_marked': False,
            'student_id': student.student_id,
            'name': student.full_name,
            'class': enrollment.class_name,
            'photo': best_match.photo,
            'message': f'Attendance marked for {student.full_name} successfully!'
        })
        
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid request body.'}, status=400)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@role_required(['admin', 'teacher'])
def holiday_list(request):
    """View to list existing manual holidays and handle creation of new ones with class applicability and date range support"""
    from .models import Holiday, Grade
    from datetime import date, timedelta
    from django.db.models import Q
    
    if request.method == 'POST':
        # Only admin is allowed to write changes (teacher is read-only)
        if request.user.profile.role != 'admin' and not request.user.is_superuser:
            messages.error(request, "Only administrators are allowed to create holidays.")
            return redirect('students:holiday_list')
            
        date_str = request.POST.get('date')
        end_date_str = request.POST.get('end_date')
        title = request.POST.get('title', '').strip()
        is_optional = request.POST.get('is_optional') == 'on'
        selected_grade_ids = request.POST.getlist('grades')
        
        if not date_str or not title:
            messages.error(request, "Start date and title are required fields.")
            return redirect('students:holiday_list')
            
        try:
            start_d = date.fromisoformat(date_str)
            end_d = date.fromisoformat(end_date_str) if end_date_str else start_d

            if end_d < start_d:
                messages.error(request, "End date cannot be earlier than start date.")
                return redirect('students:holiday_list')

            curr = start_d
            saved_count = 0
            while curr <= end_d:
                holiday_obj, created = Holiday.objects.get_or_create(
                    date=curr,
                    defaults={'title': title, 'is_optional': is_optional}
                )
                if not created:
                    holiday_obj.title = title
                    holiday_obj.is_optional = is_optional
                    holiday_obj.save()

                if selected_grade_ids:
                    holiday_obj.grades.set(selected_grade_ids)
                else:
                    holiday_obj.grades.clear()

                saved_count += 1
                curr += timedelta(days=1)

            if saved_count == 1:
                messages.success(request, f"Holiday '{title}' saved successfully!")
            else:
                messages.success(request, f"Holiday '{title}' saved for {saved_count} days ({start_d.strftime('%b %d')} - {end_d.strftime('%b %d, %Y')})!")
        except Exception as e:
            messages.error(request, f"Error creating holiday: {str(e)}")
            
        return redirect('students:holiday_list')
        
    # GET request
    holidays = Holiday.objects.all().prefetch_related('grades').order_by('-date')
    grades = Grade.objects.all().order_by('order', 'name')
    context = {
        'holidays': holidays,
        'grades': grades,
    }
    return render(request, 'students/holiday_list.html', context)


@role_required(['admin'])
@require_POST
def holiday_delete(request, pk):
    """View to delete a custom holiday"""
    from .models import Holiday
    holiday = get_object_or_404(Holiday, pk=pk)
    title = holiday.title
    date_val = holiday.date
    try:
        holiday.delete()
        messages.success(request, f"Holiday '{title}' ({date_val}) deleted successfully.")
    except Exception as e:
        messages.error(request, f"Error deleting holiday: {str(e)}")
        
    return redirect('students:holiday_list')


# ==========================================
# STUDENT CREDENTIALS MANAGEMENT (ADMIN ONLY)
# ==========================================

@role_required(['admin'])
def student_credentials_list(request):
    """
    Admin-only page to view, search, filter, manage, and reset student login credentials.
    """
    query = request.GET.get('q', '').strip()
    grade_id = request.GET.get('grade', '').strip()
    division_id = request.GET.get('division', '').strip()
    status_filter = request.GET.get('status', '').strip()

    active_year = AcademicYear.objects.filter(is_active=True).first()

    students_qs = Student.objects.filter(is_active=True).select_related(
        'user_profile',
        'user_profile__user'
    ).prefetch_related(
        'enrollments',
        'enrollments__grade',
        'enrollments__division',
        'enrollments__academic_year'
    ).order_by('first_name', 'last_name')

    # Search filter
    if query:
        students_qs = students_qs.filter(
            models.Q(student_id__icontains=query) |
            models.Q(first_name__icontains=query) |
            models.Q(last_name__icontains=query) |
            models.Q(user_profile__user__username__icontains=query)
        )

    # Grade & Division filters
    if grade_id:
        students_qs = students_qs.filter(enrollments__grade_id=grade_id)
    if division_id:
        students_qs = students_qs.filter(enrollments__division_id=division_id)

    if active_year and (grade_id or division_id):
        students_qs = students_qs.filter(enrollments__academic_year=active_year)

    students_qs = students_qs.distinct()

    all_students_list = list(students_qs)
    total_students = len(all_students_list)
    linked_count = sum(1 for s in all_students_list if hasattr(s, 'user_profile') and s.user_profile and s.user_profile.user)
    unlinked_count = total_students - linked_count

    # Sort students by Class (Grade) & Division first
    def get_sort_key(s):
        enrollment = s.current_enrollment
        if enrollment and enrollment.grade:
            g_order = getattr(enrollment.grade, 'order', 999) or 999
            g_name = enrollment.grade.name
            d_name = enrollment.division.name if enrollment.division else "No Division"
            return (g_order, g_name, d_name, s.first_name, s.last_name)
        return (9999, "ZZZ", "ZZZ", s.first_name, s.last_name)

    all_students_list.sort(key=get_sort_key)

    # Status Filter
    if status_filter == 'linked':
        all_students_list = [s for s in all_students_list if hasattr(s, 'user_profile') and s.user_profile and s.user_profile.user]
    elif status_filter == 'unlinked':
        all_students_list = [s for s in all_students_list if not (hasattr(s, 'user_profile') and s.user_profile and s.user_profile.user)]
    elif status_filter == 'active':
        all_students_list = [s for s in all_students_list if hasattr(s, 'user_profile') and s.user_profile and s.user_profile.user and s.user_profile.user.is_active]
    elif status_filter == 'inactive':
        all_students_list = [s for s in all_students_list if hasattr(s, 'user_profile') and s.user_profile and s.user_profile.user and not s.user_profile.user.is_active]

    grades = Grade.objects.all().order_by('order', 'name')
    divisions = Division.objects.all().order_by('name')

    context = {
        'students': all_students_list,
        'total_students': total_students,
        'linked_count': linked_count,
        'unlinked_count': unlinked_count,
        'grades': grades,
        'divisions': divisions,
        'query': query,
        'selected_grade': grade_id,
        'selected_division': division_id,
        'selected_status': status_filter,
    }
    return render(request, 'students/student_credentials_list.html', context)


def generate_random_password(length=5):
    """Generate a clean 5-character simple random password"""
    import secrets
    chars = "23456789abcdefghjkmnpqrstuvwxyzABCDEFGHJKMNPQRSTUVWXYZ"
    return "".join(secrets.choice(chars) for _ in range(length))


@role_required(['admin'])
@require_POST
def student_credential_create(request, student_id):
    """Create or link user credentials for a specific student"""
    student = get_object_or_404(Student, pk=student_id)
    
    if hasattr(student, 'user_profile') and student.user_profile and student.user_profile.user:
        messages.error(request, f"Student {student.full_name} already has a linked user account.")
        return redirect('students:student_credentials_list')

    username = request.POST.get('username', '').strip()
    password = request.POST.get('password', '').strip()
    password_mode = request.POST.get('password_mode', '').strip()

    if password_mode == 'random' or not password:
        password = generate_random_password(5)

    if not username:
        messages.error(request, "Username is required.")
        return redirect('students:student_credentials_list')

    if User.objects.filter(username=username).exists():
        messages.error(request, f"Username '{username}' is already taken.")
        return redirect('students:student_credentials_list')

    try:
        user = User.objects.create_user(
            username=username,
            password=password,
            first_name=student.first_name,
            last_name=student.last_name,
            email=student.email or ''
        )
        
        profile, created = UserProfile.objects.get_or_create(user=user)
        profile.role = 'student'
        profile.student_record = student
        profile.initial_password = password
        profile.save()

        messages.success(request, f"Login account created successfully for {student.full_name} (Username: {username}, Password: {password}).")
    except Exception as e:
        messages.error(request, f"Error creating credentials: {str(e)}")

    return redirect('students:student_credentials_list')


@role_required(['admin'])
@require_POST
def student_credential_reset_password(request, student_id):
    """Reset password for a student's linked user account"""
    student = get_object_or_404(Student, pk=student_id)
    
    if not (hasattr(student, 'user_profile') and student.user_profile and student.user_profile.user):
        messages.error(request, f"Student {student.full_name} does not have a linked user account.")
        return redirect('students:student_credentials_list')

    new_password = request.POST.get('new_password', '').strip()
    password_mode = request.POST.get('password_mode', '').strip()

    if password_mode == 'random' or not new_password:
        new_password = generate_random_password(5)

    user = student.user_profile.user
    user.set_password(new_password)
    user.save()

    profile = student.user_profile
    profile.initial_password = new_password
    profile.save()

    messages.success(request, f"Password successfully reset for {student.full_name} (Username: {user.username}, New Password: {new_password}).")
    return redirect('students:student_credentials_list')


@role_required(['admin'])
@require_POST
def student_credential_toggle_status(request, student_id):
    """Toggle user active / inactive status for a student account"""
    student = get_object_or_404(Student, pk=student_id)
    
    if not (hasattr(student, 'user_profile') and student.user_profile and student.user_profile.user):
        messages.error(request, f"Student {student.full_name} does not have a linked user account.")
        return redirect('students:student_credentials_list')

    user = student.user_profile.user
    user.is_active = not user.is_active
    user.save()

    status_str = "activated" if user.is_active else "deactivated"
    messages.success(request, f"Account for {student.full_name} has been {status_str}.")
    return redirect('students:student_credentials_list')


@role_required(['admin'])
@require_POST
def student_credential_bulk_create(request):
    """Bulk create credentials for unlinked students"""
    grade_id = request.POST.get('grade', '').strip()
    division_id = request.POST.get('division', '').strip()
    default_password = request.POST.get('default_password', '').strip()
    password_mode = request.POST.get('password_mode', 'default').strip()

    active_year = AcademicYear.objects.filter(is_active=True).first()

    # Query unlinked active students
    students_qs = Student.objects.filter(is_active=True, user_profile__isnull=True)

    if grade_id:
        students_qs = students_qs.filter(enrollments__grade_id=grade_id)
    if division_id:
        students_qs = students_qs.filter(enrollments__division_id=division_id)
    if active_year and (grade_id or division_id):
        students_qs = students_qs.filter(enrollments__academic_year=active_year)

    students_qs = students_qs.distinct()

    created_count = 0
    skipped_count = 0

    for student in students_qs:
        username = student.student_id.strip()
        
        if password_mode == 'random':
            pwd = generate_random_password(5)
        else:
            pwd = default_password if default_password else student.student_id.strip()

        if User.objects.filter(username=username).exists():
            skipped_count += 1
            continue

        try:
            user = User.objects.create_user(
                username=username,
                password=pwd,
                first_name=student.first_name,
                last_name=student.last_name,
                email=student.email or ''
            )
            profile, _ = UserProfile.objects.get_or_create(user=user)
            profile.role = 'student'
            profile.student_record = student
            profile.initial_password = pwd
            profile.save()
            created_count += 1
        except Exception:
            skipped_count += 1

    if created_count > 0:
        pwd_info = "with 5-character simple random passwords" if password_mode == 'random' else "with specified default password"
        messages.success(request, f"Successfully created {created_count} student user account(s) {pwd_info}!")
    if skipped_count > 0:
        messages.warning(request, f"Skipped {skipped_count} student(s) because username already exists or error occurred.")
    if created_count == 0 and skipped_count == 0:
        messages.info(request, "No unlinked students found matching your criteria.")

    return redirect('students:student_credentials_list')


@role_required(['admin'])
@require_POST
def student_credential_bulk_reset(request):
    """Bulk reset passwords to random 5-character passwords for existing accounts"""
    grade_id = request.POST.get('grade', '').strip()
    division_id = request.POST.get('division', '').strip()

    active_year = AcademicYear.objects.filter(is_active=True).first()
    students_qs = Student.objects.filter(is_active=True, user_profile__isnull=False)

    if grade_id:
        students_qs = students_qs.filter(enrollments__grade_id=grade_id)
    if division_id:
        students_qs = students_qs.filter(enrollments__division_id=division_id)
    if active_year and (grade_id or division_id):
        students_qs = students_qs.filter(enrollments__academic_year=active_year)

    students_qs = students_qs.distinct()

    reset_count = 0
    for student in students_qs:
        if hasattr(student, 'user_profile') and student.user_profile and student.user_profile.user:
            new_pwd = generate_random_password(5)
            user = student.user_profile.user
            user.set_password(new_pwd)
            user.save()
            profile = student.user_profile
            profile.initial_password = new_pwd
            profile.save()
            reset_count += 1

    messages.success(request, f"Successfully generated new 5-character random passwords for {reset_count} student accounts!")
    return redirect('students:student_credentials_list')


@role_required(['admin'])
def student_credentials_print(request):
    """Printable list of student credentials"""
    grade_id = request.GET.get('grade', '').strip()
    division_id = request.GET.get('division', '').strip()

    active_year = AcademicYear.objects.filter(is_active=True).first()

    students_qs = Student.objects.filter(is_active=True).select_related(
        'user_profile', 'user_profile__user'
    ).prefetch_related('enrollments', 'enrollments__grade', 'enrollments__division').order_by('first_name')

    if grade_id:
        students_qs = students_qs.filter(enrollments__grade_id=grade_id)
    if division_id:
        students_qs = students_qs.filter(enrollments__division_id=division_id)
    if active_year and (grade_id or division_id):
        students_qs = students_qs.filter(enrollments__academic_year=active_year)

    students_qs = students_qs.distinct()

    all_students_list = list(students_qs)
    def get_sort_key(s):
        enrollment = s.current_enrollment
        if enrollment and enrollment.grade:
            g_order = getattr(enrollment.grade, 'order', 999) or 999
            g_name = enrollment.grade.name
            d_name = enrollment.division.name if enrollment.division else "No Division"
            return (g_order, g_name, d_name, s.first_name, s.last_name)
        return (9999, "ZZZ", "ZZZ", s.first_name, s.last_name)

    all_students_list.sort(key=get_sort_key)

    grade_obj = Grade.objects.filter(id=grade_id).first() if grade_id else None
    division_obj = Division.objects.filter(id=division_id).first() if division_id else None

    context = {
        'students': all_students_list,
        'grade': grade_obj,
        'division': division_obj,
    }
    return render(request, 'students/student_credentials_print.html', context)


@role_required(['admin'])
@require_POST
def toggle_fee_maintenance(request):
    """Toggle fee section maintenance suspension for student portal"""
    from .models import GlobalSettings
    settings = GlobalSettings.load()
    settings.suspend_student_fees = not settings.suspend_student_fees
    settings.save()
    
    status_str = "SUSPENDED FOR MAINTENANCE" if settings.suspend_student_fees else "ACTIVE (Normal Mode)"
    messages.success(request, f"Student Fee Portal status updated to: {status_str}")
    
    referer = request.META.get('HTTP_REFERER')
    if referer:
        return redirect(referer)
    return redirect('students:home')


@login_required
def student_self_profile_update(request):
    """
    Dedicated view for logged-in students to update their contact details and password.
    """
    if not hasattr(request.user, 'profile') or request.user.profile.role != 'student' or not request.user.profile.student_record:
        messages.error(request, "This page is accessible to logged-in students only.")
        return redirect('students:home')

    student = request.user.profile.student_record
    user = request.user

    if request.method == 'POST':
        email = request.POST.get('email', '').strip()
        phone = request.POST.get('phone', '').strip()
        address = request.POST.get('address', '').strip()

        # Update student contact info
        student.email = email
        student.phone = phone
        student.address = address
        student.save()

        # Update user email
        if user.email != email:
            user.email = email
            user.save()

        # Handle Password Change
        current_password = request.POST.get('current_password', '').strip()
        new_password = request.POST.get('new_password', '').strip()
        confirm_password = request.POST.get('confirm_password', '').strip()

        password_changed = False

        if current_password or new_password or confirm_password:
            if not current_password:
                messages.error(request, "Current password is required to change password.")
                return redirect('students:student_self_profile_update')

            if not user.check_password(current_password):
                messages.error(request, "Current password is incorrect.")
                return redirect('students:student_self_profile_update')

            if not new_password:
                messages.error(request, "New password cannot be empty.")
                return redirect('students:student_self_profile_update')

            if len(new_password) < 4:
                messages.error(request, "New password must be at least 4 characters long.")
                return redirect('students:student_self_profile_update')

            if new_password != confirm_password:
                messages.error(request, "New password and Confirm password do not match.")
                return redirect('students:student_self_profile_update')

            # Apply new password
            user.set_password(new_password)
            user.save()

            # Save initial_password in profile for reference
            profile = user.profile
            profile.initial_password = new_password
            profile.save()

            # Prevent logout session drop
            from django.contrib.auth import update_session_auth_hash
            update_session_auth_hash(request, user)
            password_changed = True

        if password_changed:
            messages.success(request, "Your contact details and password have been updated successfully!")
        else:
            messages.success(request, "Your contact details have been updated successfully!")

        return redirect('students:student_profile', pk=student.id)

    enrollment = student.current_enrollment

    context = {
        'student': student,
        'user': user,
        'enrollment': enrollment,
    }
    return render(request, 'students/student_self_profile_update.html', context)







